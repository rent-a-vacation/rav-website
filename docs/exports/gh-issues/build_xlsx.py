"""Generate a branded Excel export of all open GitHub issues.

Reads `_issues-raw.json` (produced by `gh issue list ... --json ...`) and
writes `rav-open-issues-<YYYY-MM-DD>.xlsx` in the same folder.

Brand colors come from docs/brand-assets/BRAND-STYLE-GUIDE.md.
"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).resolve().parent
RAW = HERE / "_issues-raw.json"
DISCUSSIONS_RAW = HERE / "_discussions-raw.json"
TODAY = datetime.now(timezone.utc).strftime("%Y-%m-%d")
OUT = HERE / f"rav-open-issues-{TODAY}.xlsx"

# ── Brand palette (BRAND-STYLE-GUIDE.md) ────────────────────────────────────
TEAL = "1C7268"
CORAL = "E8703A"
CREAM = "F8F6F3"
NAVY = "1D2E38"
SAND = "F0EBE3"
SLATE = "6B7B85"
SUCCESS = "1FA66E"
WARNING = "F59E0B"
ERROR = "E53E3E"
WHITE = "FFFFFF"

# Label → tier classification
TIER_PRIORITY = ["blocked", "pre-launch", "post-launch", "post-beta", "future"]
TYPE_LABELS = {"bug", "enhancement", "docs", "refactor", "idea"}
AREA_LABELS = {"marketplace", "platform", "experience", "vacation-wishes"}


def classify(labels: list[str]) -> tuple[str, str, str]:
    """Return (priority, type, area) derived from label set."""
    lset = {l.lower() for l in labels}
    priority = next((p for p in TIER_PRIORITY if p in lset), "unscheduled")
    issue_type = next((t for t in ["bug", "enhancement", "docs", "refactor", "idea"] if t in lset), "")
    areas = sorted(a for a in AREA_LABELS if a in lset)
    area = ", ".join(areas)
    return priority, issue_type, area


def truncate(s: str, n: int = 500) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s if len(s) <= n else s[: n - 1] + "…"


def fmt_date(iso: str) -> str:
    if not iso:
        return ""
    return datetime.fromisoformat(iso.replace("Z", "+00:00")).strftime("%Y-%m-%d")


# ── Load issues ─────────────────────────────────────────────────────────────
with RAW.open(encoding="utf-8") as fh:
    issues = json.load(fh)

issues.sort(
    key=lambda i: (
        TIER_PRIORITY.index(classify([l["name"] for l in i["labels"]])[0])
        if classify([l["name"] for l in i["labels"]])[0] in TIER_PRIORITY
        else len(TIER_PRIORITY),
        -i["number"],
    )
)

# ── Workbook ────────────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Open Issues ────────────────────────────────────────────────────
ws = wb.active
ws.title = "Open Issues"

title_fill = PatternFill("solid", fgColor=TEAL)
title_font = Font(name="Calibri", size=18, bold=True, color=WHITE)
subtitle_fill = PatternFill("solid", fgColor=CORAL)
subtitle_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
header_fill = PatternFill("solid", fgColor=NAVY)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
band_fill = PatternFill("solid", fgColor=CREAM)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="D5D8DC")
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

columns = [
    ("#", 8),
    ("Title", 55),
    ("Priority", 14),
    ("Type", 14),
    ("Area", 22),
    ("Milestone", 28),
    ("Labels", 32),
    ("Assignees", 18),
    ("Author", 14),
    ("Comments", 10),
    ("Created", 12),
    ("Updated", 12),
    ("URL", 48),
    ("Summary", 90),
]

# Row 1: report title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
c = ws.cell(row=1, column=1, value="Rent-A-Vacation — Open GitHub Issues")
c.fill = title_fill
c.font = title_font
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 36

# Row 2: subtitle
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
c = ws.cell(
    row=2,
    column=1,
    value=f"Repo: rent-a-vacation/rav-website  ·  Generated {TODAY}  ·  {len(issues)} open issues",
)
c.fill = subtitle_fill
c.font = subtitle_font
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 22

# Row 3: column headers
HEADER_ROW = 3
for idx, (name, width) in enumerate(columns, start=1):
    cell = ws.cell(row=HEADER_ROW, column=idx, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = cell_border
    ws.column_dimensions[get_column_letter(idx)].width = width
ws.row_dimensions[HEADER_ROW].height = 24

# Priority chip colors
priority_fill = {
    "blocked": PatternFill("solid", fgColor=ERROR),
    "pre-launch": PatternFill("solid", fgColor=CORAL),
    "post-launch": PatternFill("solid", fgColor=TEAL),
    "post-beta": PatternFill("solid", fgColor=WARNING),
    "future": PatternFill("solid", fgColor=SAND),
    "unscheduled": PatternFill("solid", fgColor="EAE8E4"),
}
priority_font = {
    "blocked": Font(bold=True, color=WHITE),
    "pre-launch": Font(bold=True, color=WHITE),
    "post-launch": Font(bold=True, color=WHITE),
    "post-beta": Font(bold=True, color=NAVY),
    "future": Font(bold=False, color=NAVY),
    "unscheduled": Font(bold=False, color=NAVY),
}

# Issue rows
for row_idx, issue in enumerate(issues, start=HEADER_ROW + 1):
    labels = [l["name"] for l in issue["labels"]]
    priority, issue_type, area = classify(labels)
    assignees = ", ".join(a["login"] for a in issue["assignees"])
    milestone = issue["milestone"]["title"] if issue["milestone"] else ""

    values = [
        issue["number"],
        issue["title"],
        priority,
        issue_type,
        area,
        milestone,
        ", ".join(sorted(labels)),
        assignees,
        issue["author"]["login"],
        len(issue["comments"]),
        fmt_date(issue["createdAt"]),
        fmt_date(issue["updatedAt"]),
        issue["url"],
        truncate(issue["body"], 500),
    ]

    band = (row_idx - HEADER_ROW) % 2 == 0  # alternating band on even data rows
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = cell_border
        if band:
            cell.fill = band_fill
        cell.alignment = left if col_idx in (2, 7, 13, 14) else center

    # Priority chip
    pcell = ws.cell(row=row_idx, column=3)
    pcell.fill = priority_fill[priority]
    pcell.font = priority_font[priority]
    pcell.alignment = center

    # Hyperlink on URL column
    url_cell = ws.cell(row=row_idx, column=13)
    url_cell.hyperlink = issue["url"]
    url_cell.font = Font(color=TEAL, underline="single")

    # Issue # also hyperlinked
    num_cell = ws.cell(row=row_idx, column=1)
    num_cell.hyperlink = issue["url"]
    num_cell.font = Font(color=TEAL, bold=True, underline="single")

# Freeze panes below header, autofilter on the header row
ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)
last_col = get_column_letter(len(columns))
ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{HEADER_ROW + len(issues)}"

# Sensible row height for wrapped text
for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + len(issues)):
    ws.row_dimensions[r].height = 60

# ── Sheet 2: Summary ────────────────────────────────────────────────────────
sw = wb.create_sheet("Summary")

sw.merge_cells("A1:D1")
c = sw.cell(row=1, column=1, value="Open-issue rollups")
c.fill = title_fill
c.font = title_font
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
sw.row_dimensions[1].height = 32

def write_table(start_row: int, heading: str, rows: list[tuple[str, int]]) -> int:
    sw.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    h = sw.cell(row=start_row, column=1, value=heading)
    h.fill = subtitle_fill
    h.font = subtitle_font
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sw.row_dimensions[start_row].height = 22

    sw.cell(row=start_row + 1, column=1, value="Bucket").font = header_font
    sw.cell(row=start_row + 1, column=1).fill = header_fill
    sw.cell(row=start_row + 1, column=1).alignment = center
    sw.cell(row=start_row + 1, column=2, value="Open issues").font = header_font
    sw.cell(row=start_row + 1, column=2).fill = header_fill
    sw.cell(row=start_row + 1, column=2).alignment = center

    r = start_row + 2
    for label, count in rows:
        a = sw.cell(row=r, column=1, value=label)
        b = sw.cell(row=r, column=2, value=count)
        for cc in (a, b):
            cc.border = cell_border
        a.alignment = left
        b.alignment = center
        r += 1
    return r + 1  # gap row

# Counts
def count_by(key_fn) -> list[tuple[str, int]]:
    counts: dict[str, int] = {}
    for i in issues:
        for k in key_fn(i):
            counts[k] = counts.get(k, 0) + 1
    return sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))

priority_counts = count_by(lambda i: [classify([l["name"] for l in i["labels"]])[0]])
type_counts = count_by(lambda i: [classify([l["name"] for l in i["labels"]])[1] or "(unset)"])
area_counts = count_by(lambda i: [classify([l["name"] for l in i["labels"]])[2] or "(unset)"])
milestone_counts = count_by(lambda i: [i["milestone"]["title"] if i["milestone"] else "(no milestone)"])
label_counts = count_by(lambda i: [l["name"] for l in i["labels"]])

sw.column_dimensions["A"].width = 38
sw.column_dimensions["B"].width = 14

next_row = 3
next_row = write_table(next_row, "By priority", priority_counts)
next_row = write_table(next_row, "By type", type_counts)
next_row = write_table(next_row, "By area", area_counts)
next_row = write_table(next_row, "By milestone", milestone_counts)
next_row = write_table(next_row, "By label", label_counts)

# ── Sheet 3: Legend ─────────────────────────────────────────────────────────
lg = wb.create_sheet("Legend")
lg.merge_cells("A1:B1")
c = lg.cell(row=1, column=1, value="Priority chip legend")
c.fill = title_fill
c.font = title_font
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
lg.row_dimensions[1].height = 32

legend_rows = [
    ("blocked", "Has the `blocked` label — waiting on an external dependency."),
    ("pre-launch", "Required before public launch."),
    ("post-launch", "Can wait until after launch."),
    ("post-beta", "Targeted for after a beta cohort."),
    ("future", "Long-horizon idea or backlog item."),
    ("unscheduled", "No tier label set — needs triage."),
]
lg.cell(row=2, column=1, value="Chip").fill = header_fill
lg.cell(row=2, column=1).font = header_font
lg.cell(row=2, column=1).alignment = center
lg.cell(row=2, column=2, value="Meaning").fill = header_fill
lg.cell(row=2, column=2).font = header_font
lg.cell(row=2, column=2).alignment = center
lg.column_dimensions["A"].width = 18
lg.column_dimensions["B"].width = 80

for idx, (key, desc) in enumerate(legend_rows, start=3):
    a = lg.cell(row=idx, column=1, value=key)
    b = lg.cell(row=idx, column=2, value=desc)
    a.fill = priority_fill[key]
    a.font = priority_font[key]
    a.alignment = center
    b.alignment = left
    b.border = cell_border
    a.border = cell_border

# ── Sheets 4 & 5: Discussions ───────────────────────────────────────────────
# Source = GitHub Discussions area (https://github.com/rent-a-vacation/rav-website/discussions)
# Used by the team as a TODO / Action-Items board.
discussions = []
if DISCUSSIONS_RAW.exists():
    with DISCUSSIONS_RAW.open(encoding="utf-8") as fh:
        discussions = json.load(fh)["data"]["repository"]["discussions"]["nodes"]


def comment_total(d: dict) -> int:
    n = d["comments"].get("totalCount", len(d["comments"]["nodes"]))
    n += sum(len(c["replies"]["nodes"]) for c in d["comments"]["nodes"])
    return n


def last_activity(d: dict) -> str:
    iso = d["updatedAt"]
    for c in d["comments"]["nodes"]:
        if c["createdAt"] > iso:
            iso = c["createdAt"]
        for r in c["replies"]["nodes"]:
            if r["createdAt"] > iso:
                iso = r["createdAt"]
    return fmt_date(iso)


# Sheet 4: Action Items (Discussions) — one row per discussion
ai = wb.create_sheet("Action Items (Discussions)")

ai_columns = [
    ("#", 8),
    ("Title", 55),
    ("Category", 18),
    ("Author", 16),
    ("Comments", 10),
    ("Created", 12),
    ("Last activity", 14),
    ("URL", 48),
    ("Body excerpt", 90),
]

ai.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ai_columns))
c = ai.cell(row=1, column=1, value="Action Items — sourced from the GitHub Discussions area")
c.fill = title_fill
c.font = title_font
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ai.row_dimensions[1].height = 36

ai.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ai_columns))
c = ai.cell(
    row=2,
    column=1,
    value=(
        f"Source: github.com/rent-a-vacation/rav-website/discussions  ·  "
        f"Generated {TODAY}  ·  {len(discussions)} open discussions  ·  "
        f"Each row is one Discussion the team is using as a TODO / action item"
    ),
)
c.fill = subtitle_fill
c.font = subtitle_font
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ai.row_dimensions[2].height = 22

for idx, (name, width) in enumerate(ai_columns, start=1):
    cell = ai.cell(row=3, column=idx, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = cell_border
    ai.column_dimensions[get_column_letter(idx)].width = width
ai.row_dimensions[3].height = 24

category_fill = {
    "Announcements": PatternFill("solid", fgColor=CORAL),
    "General": PatternFill("solid", fgColor=TEAL),
    "Ideas": PatternFill("solid", fgColor=WARNING),
    "Polls": PatternFill("solid", fgColor=SAND),
    "Q&A": PatternFill("solid", fgColor=SUCCESS),
    "Show and tell": PatternFill("solid", fgColor=NAVY),
}
category_font = {
    "Announcements": Font(bold=True, color=WHITE),
    "General": Font(bold=True, color=WHITE),
    "Ideas": Font(bold=True, color=NAVY),
    "Polls": Font(bold=True, color=NAVY),
    "Q&A": Font(bold=True, color=WHITE),
    "Show and tell": Font(bold=True, color=WHITE),
}

discussions_sorted = sorted(discussions, key=lambda d: d["updatedAt"], reverse=True)

for row_idx, d in enumerate(discussions_sorted, start=4):
    cat = d["category"]["name"]
    values = [
        d["number"],
        d["title"],
        cat,
        (d["author"] or {}).get("login", ""),
        comment_total(d),
        fmt_date(d["createdAt"]),
        last_activity(d),
        d["url"],
        truncate(d["body"], 700),
    ]

    band = (row_idx - 3) % 2 == 0
    for col_idx, value in enumerate(values, start=1):
        cell = ai.cell(row=row_idx, column=col_idx, value=value)
        cell.border = cell_border
        if band:
            cell.fill = band_fill
        cell.alignment = left if col_idx in (2, 8, 9) else center

    # Category chip
    cat_cell = ai.cell(row=row_idx, column=3)
    cat_cell.fill = category_fill.get(cat, PatternFill("solid", fgColor="EAE8E4"))
    cat_cell.font = category_font.get(cat, Font(color=NAVY))
    cat_cell.alignment = center

    url_cell = ai.cell(row=row_idx, column=8)
    url_cell.hyperlink = d["url"]
    url_cell.font = Font(color=TEAL, underline="single")

    num_cell = ai.cell(row=row_idx, column=1)
    num_cell.hyperlink = d["url"]
    num_cell.font = Font(color=TEAL, bold=True, underline="single")

    ai.row_dimensions[row_idx].height = 70

if discussions_sorted:
    ai.freeze_panes = ai.cell(row=4, column=1)
    last_col_ai = get_column_letter(len(ai_columns))
    ai.auto_filter.ref = f"A3:{last_col_ai}{3 + len(discussions_sorted)}"


# Sheet 5: Open Checkboxes (parsed from Discussion bodies + comments + replies)
cb = wb.create_sheet("Open Checkboxes (Discussions)")

cb_columns = [
    ("Discussion #", 12),
    ("Discussion title", 45),
    ("Category", 18),
    ("Action item (open checkbox)", 70),
    ("Found in", 18),
    ("Posted by", 16),
    ("Posted at", 12),
    ("URL", 48),
]

cb.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cb_columns))
c = cb.cell(row=1, column=1, value="Open checkboxes — extracted from GitHub Discussions")
c.fill = title_fill
c.font = title_font
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
cb.row_dimensions[1].height = 36

cb.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cb_columns))
c = cb.cell(
    row=2,
    column=1,
    value=(
        "Each row is a single unchecked `- [ ]` line found inside a Discussion's body, "
        "comment, or comment reply. Done items (`- [x]`) are excluded."
    ),
)
c.fill = subtitle_fill
c.font = subtitle_font
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
cb.row_dimensions[2].height = 22

for idx, (name, width) in enumerate(cb_columns, start=1):
    cell = cb.cell(row=3, column=idx, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = cell_border
    cb.column_dimensions[get_column_letter(idx)].width = width
cb.row_dimensions[3].height = 24

CHECKBOX_RE = re.compile(r"^\s*[-*]\s*\[\s\]\s+(.+?)\s*$", re.MULTILINE)


def extract_checkboxes(text: str) -> list[str]:
    if not text:
        return []
    return [m.group(1).strip() for m in CHECKBOX_RE.finditer(text)]


checkbox_rows: list[tuple] = []  # (disc_num, title, cat, item, where, who, when, url)
for d in discussions_sorted:
    disc_num = d["number"]
    disc_title = d["title"]
    cat = d["category"]["name"]
    url = d["url"]

    for item in extract_checkboxes(d.get("body", "")):
        checkbox_rows.append(
            (disc_num, disc_title, cat, item, "Body", (d["author"] or {}).get("login", ""), fmt_date(d["createdAt"]), url)
        )
    for c in d["comments"]["nodes"]:
        for item in extract_checkboxes(c.get("body", "")):
            checkbox_rows.append(
                (disc_num, disc_title, cat, item, "Comment", (c["author"] or {}).get("login", ""), fmt_date(c["createdAt"]), url)
            )
        for r in c["replies"]["nodes"]:
            for item in extract_checkboxes(r.get("body", "")):
                checkbox_rows.append(
                    (disc_num, disc_title, cat, item, "Reply", (r["author"] or {}).get("login", ""), fmt_date(r["createdAt"]), url)
                )

if not checkbox_rows:
    cb.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(cb_columns))
    c = cb.cell(
        row=4,
        column=1,
        value="No open checkboxes found in any Discussion body, comment, or reply. (Done items are excluded.)",
    )
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.font = Font(italic=True, color=SLATE)
    cb.row_dimensions[4].height = 28
else:
    for row_idx, row in enumerate(checkbox_rows, start=4):
        band = (row_idx - 3) % 2 == 0
        for col_idx, value in enumerate(row, start=1):
            cell = cb.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            if band:
                cell.fill = band_fill
            cell.alignment = left if col_idx in (2, 4, 8) else center

        cat_cell = cb.cell(row=row_idx, column=3)
        cat_cell.fill = category_fill.get(row[2], PatternFill("solid", fgColor="EAE8E4"))
        cat_cell.font = category_font.get(row[2], Font(color=NAVY))
        cat_cell.alignment = center

        url_cell = cb.cell(row=row_idx, column=8)
        url_cell.hyperlink = row[7]
        url_cell.font = Font(color=TEAL, underline="single")

        num_cell = cb.cell(row=row_idx, column=1)
        num_cell.hyperlink = row[7]
        num_cell.font = Font(color=TEAL, bold=True, underline="single")

        cb.row_dimensions[row_idx].height = 36

    cb.freeze_panes = cb.cell(row=4, column=1)
    last_col_cb = get_column_letter(len(cb_columns))
    cb.auto_filter.ref = f"A3:{last_col_cb}{3 + len(checkbox_rows)}"

# Reorder so the meeting-friendly sheets come right after Open Issues
desired_order = [
    "Open Issues",
    "Action Items (Discussions)",
    "Open Checkboxes (Discussions)",
    "Summary",
    "Legend",
]
wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]

try:
    wb.save(OUT)
    final_path = OUT
except PermissionError:
    stamp = datetime.now(timezone.utc).strftime("%H%M%S")
    final_path = HERE / f"rav-open-issues-{TODAY}-{stamp}.xlsx"
    wb.save(final_path)
    print(f"NOTE: {OUT.name} was locked (open in Excel?) — wrote {final_path.name} instead")

print(
    f"Wrote {final_path}  ({len(issues)} issues, "
    f"{len(discussions_sorted)} discussions, {len(checkbox_rows)} open checkboxes)"
)

"""
Generate RAV-branded QA Scenario Test Workbook
Creates a professional Excel workbook for tracking E2E scenario test runs.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from copy import copy
import os
from datetime import datetime

# ── RAV Brand Colors ──
TEAL = "0D6B5C"
TEAL_LIGHT = "E6F2EF"
TEAL_MED = "B3DDD3"
CORAL = "E86A4A"
CORAL_LIGHT = "FDE8E2"
DARK = "1A1A2E"
WHITE = "FFFFFF"
GRAY_BG = "F8F9FA"
GRAY_BORDER = "DEE2E6"
GRAY_TEXT = "6C757D"
AMBER = "F59E0B"
AMBER_LIGHT = "FEF3C7"
RED = "DC2626"
RED_LIGHT = "FEE2E2"
GREEN = "16A34A"
GREEN_LIGHT = "DCFCE7"
BLUE = "2563EB"
BLUE_LIGHT = "DBEAFE"

# ── Reusable Styles ──
thin_border = Border(
    left=Side(style='thin', color=GRAY_BORDER),
    right=Side(style='thin', color=GRAY_BORDER),
    top=Side(style='thin', color=GRAY_BORDER),
    bottom=Side(style='thin', color=GRAY_BORDER),
)

header_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type='solid')
header_font = Font(name='Calibri', bold=True, color=WHITE, size=11)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

subheader_fill = PatternFill(start_color=TEAL_LIGHT, end_color=TEAL_LIGHT, fill_type='solid')
subheader_font = Font(name='Calibri', bold=True, color=TEAL, size=11)

body_font = Font(name='Calibri', size=11, color=DARK)
body_align = Alignment(vertical='center', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

title_font = Font(name='Calibri', bold=True, color=TEAL, size=18)
subtitle_font = Font(name='Calibri', bold=True, color=DARK, size=14)
section_font = Font(name='Calibri', bold=True, color=TEAL, size=13)
kpi_value_font = Font(name='Calibri', bold=True, color=TEAL, size=28)
kpi_label_font = Font(name='Calibri', color=GRAY_TEXT, size=10)

pass_fill = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type='solid')
fail_fill = PatternFill(start_color=RED_LIGHT, end_color=RED_LIGHT, fill_type='solid')
blocked_fill = PatternFill(start_color=AMBER_LIGHT, end_color=AMBER_LIGHT, fill_type='solid')
skip_fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid')


def style_header_row(ws, row, max_col, fill=None, font=None):
    """Apply header styling to a row."""
    f = fill or header_fill
    fn = font or header_font
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = f
        cell.font = fn
        cell.alignment = header_align
        cell.border = thin_border


def style_data_cell(ws, row, col, align=None):
    """Apply standard data cell styling."""
    cell = ws.cell(row=row, column=col)
    cell.font = body_font
    cell.alignment = align or body_align
    cell.border = thin_border
    return cell


def add_status_validation(ws, cell_range):
    """Add dropdown validation for Pass/Fail/Blocked/Skip."""
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"PASS,FAIL,BLOCKED,SKIP,N/A"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Please select: PASS, FAIL, BLOCKED, SKIP, or N/A"
    dv.errorTitle = "Invalid Status"
    dv.prompt = "Select test result"
    dv.promptTitle = "Status"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_step_status_validation(ws, cell_range):
    """Add dropdown for step-level pass/fail."""
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"\u2705,\u274C,\u23ED,\u26A0"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.prompt = "\u2705=Pass, \u274C=Fail, \u23ED=Skip, \u26A0=Blocked"
    dv.promptTitle = "Step Result"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_conditional_formatting(ws, cell_range):
    """Add color-coding for PASS/FAIL/BLOCKED/SKIP."""
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator='equal', formula=['"PASS"'], fill=pass_fill,
        font=Font(color=GREEN, bold=True)
    ))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator='equal', formula=['"FAIL"'], fill=fail_fill,
        font=Font(color=RED, bold=True)
    ))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator='equal', formula=['"BLOCKED"'], fill=blocked_fill,
        font=Font(color=AMBER, bold=True)
    ))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator='equal', formula=['"SKIP"'], fill=skip_fill,
        font=Font(color=GRAY_TEXT)
    ))


def add_step_conditional_formatting(ws, cell_range):
    """Add color-coding for emoji step results."""
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator='equal', formula=['"\u2705"'], fill=pass_fill
    ))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator='equal', formula=['"\u274C"'], fill=fail_fill
    ))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator='equal', formula=['"\u23ED"'], fill=skip_fill
    ))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator='equal', formula=['"\u26A0"'], fill=blocked_fill
    ))


# ══════════════════════════════════════════════════════════════
# SCENARIO DATA
# ══════════════════════════════════════════════════════════════

SCENARIOS = [
    {
        "id": "S-01", "name": "New User Onboarding & Approval",
        "tier": "Foundation", "est_time": "10 min",
        "roles": "New User, RAV Admin",
        "capability": "Signup \u2192 Approval \u2192 Access",
        "preconditions": "Use an email not in the system (e.g., testuser-[date]@gmail.com)",
        "steps": [
            ("New User", "Open incognito \u2192 /signup", "/signup", "Signup form loads with email + password + Google OAuth"),
            ("New User", "Enter name, email, password \u2192 Submit", "/signup", "Success message, redirected to /pending-approval"),
            ("New User", "Try navigating to /rentals", "/rentals", "Blocked \u2014 redirected to /pending-approval"),
            ("Admin", "Log in as dev-admin@rent-a-vacation.com", "/login", "Dashboard loads"),
            ("Admin", "Navigate to /admin?tab=pending-approvals", "/admin", "New user in pending list"),
            ("Admin", "Click Approve on the new user", "/admin", "Status changes to approved"),
            ("New User", "Refresh or re-login", "\u2014", "Redirected to /rentals \u2014 full access"),
            ("New User", "Browse /rentals, click a listing", "/property/:id", "Property details load"),
        ]
    },
    {
        "id": "S-02", "name": "Owner Role Upgrade & Verification",
        "tier": "Admin & Platform", "est_time": "10 min",
        "roles": "Renter 4, RAV Admin",
        "capability": "Role upgrade \u2192 Doc upload \u2192 Admin verify",
        "preconditions": "Renter 4 (renter004@) has approved renter account, no owner role",
        "steps": [
            ("Renter 4", "Log in \u2192 /owner-dashboard", "/owner-dashboard", "RoleUpgradeDialog appears"),
            ("Renter 4", "Submit role upgrade request", "/owner-dashboard", "Toast: Request submitted"),
            ("Admin", "Log in \u2192 /admin?tab=pending-approvals", "/admin", "Role upgrade request visible"),
            ("Admin", "Approve role upgrade", "/admin", "Renter 4 now has property_owner role"),
            ("Renter 4", "Refresh page or re-login", "/owner-dashboard", "Owner dashboard loads (no upgrade dialog)"),
            ("Renter 4", "Account tab \u2192 Verification section", "/owner-dashboard?tab=account", "Upload fields visible"),
            ("Renter 4", "Upload verification documents", "/owner-dashboard?tab=account", "Status: Pending verification"),
            ("Admin", "/admin?tab=verifications", "/admin", "Docs in verification queue"),
            ("Admin", "Review and approve verification", "/admin", "Status: Verified"),
            ("Renter 4", "Refresh Account tab", "/owner-dashboard?tab=account", "Verification badge: Verified"),
        ]
    },
    {
        "id": "S-03", "name": "Property Listing Creation & Approval",
        "tier": "Foundation", "est_time": "12 min",
        "roles": "Owner 1 (Alex, Pro), RAV Admin",
        "capability": "List property \u2192 Admin review \u2192 Goes live",
        "preconditions": "Owner 1 logged in, not at listing limit (Pro = 10)",
        "steps": [
            ("Owner 1", "Navigate to /list-property", "/list-property", "3-step form loads"),
            ("Owner 1", "Step 1: Fill property details (Las Vegas, Hilton, 2BR/2BA, sleeps 6)", "/list-property", "Fields accept input, amenities work"),
            ("Owner 1", "Click Next", "/list-property", "Step 2: Dates & Pricing"),
            ("Owner 1", "Step 2: Check-in 60d out, 7 nights, $200/night, Moderate cancel, Open for Bidding", "/list-property", "Price: $1,400 owner + $210 markup = $1,610"),
            ("Owner 1", "Click Next", "/list-property", "Step 3: Review & Submit"),
            ("Owner 1", "Review \u2192 Submit for Review", "/list-property", "Redirected to /owner-dashboard?tab=my-listings"),
            ("Owner 1", "Check My Listings tab", "/owner-dashboard", "Status: Pending Approval"),
            ("Admin", "Log in \u2192 /admin?tab=listings", "/admin", "New listing in pending queue"),
            ("Admin", "Review \u2192 Approve", "/admin", "Status: Active"),
            ("Owner 1", "Refresh My Listings", "/owner-dashboard", "Status: Active"),
            ("Renter 1", "Log in \u2192 /rentals \u2192 search Las Vegas", "/rentals", "New listing appears"),
            ("Renter 1", "Click listing", "/property/:id", "Full details visible with bidding badge"),
        ]
    },
    {
        "id": "S-04", "name": "Direct Booking \u2014 Happy Path",
        "tier": "Core Marketplace", "est_time": "15 min",
        "roles": "Renter 1, Owner 1, RAV Admin",
        "capability": "Browse \u2192 Book \u2192 Pay \u2192 Owner confirm \u2192 Escrow",
        "preconditions": "Active listing from Owner 1 with direct booking (non-bidding)",
        "steps": [
            ("Renter 1", "Log in \u2192 /rentals \u2192 find Owner 1 fixed-price listing", "/rentals", "Listing visible with price"),
            ("Renter 1", "Click listing \u2192 view details", "/property/:id", "Price breakdown, cancellation policy detail"),
            ("Renter 1", "Click Book Now", "/checkout", "Checkout with fee breakdown, Stripe form"),
            ("Renter 1", "Pay with test card 4242...4242", "/checkout", "Redirect to /booking-success"),
            ("Renter 1", "View booking confirmation", "/booking-success", "Timeline: Payment \u2705, Owner Confirm \u23f3"),
            ("Owner 1", "Log in \u2192 check notification bell", "Any page", "Notification: New booking"),
            ("Owner 1", "/owner-dashboard?tab=bookings-earnings", "/owner-dashboard", "Booking with Confirm button + countdown"),
            ("Owner 1", "Confirm \u2192 enter resort confirmation RST123456", "/owner-dashboard", "Status: Confirmed"),
            ("Renter 1", "/my-trips?tab=bookings", "/my-trips", "Booking: Confirmed, timeline updated"),
            ("Admin", "/admin?tab=escrow", "/admin", "Escrow: Pending Verification"),
            ("Admin", "Verify confirmation \u2192 Release escrow", "/admin", "Escrow: Released"),
            ("Owner 1", "Check Bookings & Earnings", "/owner-dashboard", "Earnings updated, payout visible"),
        ]
    },
    {
        "id": "S-05", "name": "Bidding & Counter-Offer Workflow",
        "tier": "Core Marketplace", "est_time": "15 min",
        "roles": "Renter 1, Owner 1",
        "capability": "Bid \u2192 Counter \u2192 Accept \u2192 Booking",
        "preconditions": "Active listing from Owner 1 with open_for_bidding = true",
        "steps": [
            ("Renter 1", "/rentals \u2192 find bidding listing from Owner 1", "/rentals", "Open for Bidding badge visible"),
            ("Renter 1", "Click listing \u2192 Place Bid", "/property/:id", "BidFormDialog opens"),
            ("Renter 1", "Bid at 80% of list price, 4 guests, add message \u2192 Submit", "/property/:id", "Toast: Bid submitted"),
            ("Renter 1", "/my-trips?tab=offers", "/my-trips", "Bid: Pending"),
            ("Owner 1", "Log in \u2192 check notification bell", "Any", "Notification: New bid received"),
            ("Owner 1", "My Listings \u2192 View Bids (BidsManagerDialog)", "/owner-dashboard", "Renter 1's bid visible"),
            ("Owner 1", "Counter at 90% of list price with message \u2192 Submit", "/owner-dashboard", "Counter-offer sent"),
            ("Renter 1", "Check /my-trips?tab=offers", "/my-trips", "Counter-offer amount + message shown"),
            ("Renter 1", "Accept counter-offer", "/my-trips", "Redirected to /checkout at counter amount"),
            ("Renter 1", "Complete Stripe payment", "/booking-success", "Booking created at counter price"),
            ("Owner 1", "Check Bookings & Earnings", "/owner-dashboard", "New booking from bid visible"),
        ]
    },
    {
        "id": "S-06", "name": "Travel Request & Owner Proposal",
        "tier": "Core Marketplace", "est_time": "15 min",
        "roles": "Renter 2 (Liam), Owner 2 (Maria, Marriott)",
        "capability": "Post request \u2192 Owner proposes \u2192 Accept \u2192 Book",
        "preconditions": "Owner 2 has active Marriott listings in Orlando",
        "steps": [
            ("Renter 2", "Log in \u2192 /bidding", "/bidding", "TravelRequestForm loads"),
            ("Renter 2", "Fill: Orlando FL, 45d out, 5 nights, 4 guests, $1000-2000, flexible \u00b13d \u2192 Submit", "/bidding", "Toast: Travel request posted"),
            ("Owner 2", "Log in \u2192 /owner-dashboard \u2192 travel requests section", "/owner-dashboard", "Renter 2's matched request visible"),
            ("Owner 2", "View request \u2192 Submit Proposal", "/owner-dashboard", "Proposal form opens"),
            ("Owner 2", "Select Grande Vista, $1,500, matching dates, add message \u2192 Submit", "/owner-dashboard", "Toast: Proposal sent"),
            ("Renter 2", "Check notification bell", "Any", "Notification: Owner proposal received"),
            ("Renter 2", "/my-trips?tab=offers", "/my-trips", "Proposal with property, price, dates"),
            ("Renter 2", "Accept Proposal", "/my-trips", "Redirected to /checkout"),
            ("Renter 2", "Complete Stripe payment", "/booking-success", "Booking at $1,500"),
            ("Owner 2", "Check bookings", "/owner-dashboard", "Booking from travel request visible"),
        ]
    },
    {
        "id": "S-07", "name": "Pre-Booking Inquiry & Messaging",
        "tier": "Communication", "est_time": "8 min",
        "roles": "Renter 3 (Olivia, Premium), Owner 3 (James, Disney)",
        "capability": "Ask question \u2192 Owner replies \u2192 Conversation",
        "preconditions": "Owner 3 has active Disney listing",
        "steps": [
            ("Renter 3", "Log in \u2192 find Disney listing \u2192 /property/:id", "/property/:id", "Ask the Owner button visible"),
            ("Renter 3", "Click Ask the Owner \u2192 type question \u2192 Send", "/property/:id", "Toast: Message sent"),
            ("Owner 3", "Log in \u2192 check notification bell", "Any", "Notification: New inquiry"),
            ("Owner 3", "Click notification \u2192 view InquiryThread", "InquiryThread", "Renter's message visible"),
            ("Owner 3", "Reply with answer \u2192 Send", "InquiryThread", "Reply sent"),
            ("Renter 3", "Check notification \u2192 view thread", "InquiryThread", "Owner's reply visible"),
            ("Renter 3", "Reply again \u2192 Send", "InquiryThread", "Conversation continues"),
        ]
    },
    {
        "id": "S-08", "name": "Booking Messaging & Realtime",
        "tier": "Communication", "est_time": "8 min",
        "roles": "Renter 1, Owner 1",
        "capability": "Post-booking thread \u2192 Realtime delivery",
        "preconditions": "Confirmed booking between Renter 1 and Owner 1",
        "steps": [
            ("Renter 1", "/my-trips?tab=bookings \u2192 select booking \u2192 open messages", "/my-trips", "BookingMessageThread loads"),
            ("Renter 1", "Type message: 'What's the WiFi password?' \u2192 Send", "/my-trips", "Message appears instantly"),
            ("Owner 1", "Log in (separate browser) \u2192 check bell", "Any", "Notification arrives (realtime, no refresh)"),
            ("Owner 1", "Navigate to booking \u2192 open messages", "/owner-dashboard", "Renter's message visible"),
            ("Owner 1", "Reply with WiFi info \u2192 Send", "/owner-dashboard", "Message sent"),
            ("Renter 1", "Check thread WITHOUT refreshing", "/my-trips", "Owner's reply appears in realtime"),
        ]
    },
    {
        "id": "S-09", "name": "Renter Cancellation Flow",
        "tier": "Communication", "est_time": "10 min",
        "roles": "Renter 1, Owner 1, RAV Admin",
        "capability": "Cancel \u2192 Policy refund \u2192 Status updates",
        "preconditions": "Confirmed booking between Renter 1 and Owner 1",
        "steps": [
            ("Renter 1", "/my-trips?tab=bookings \u2192 select booking", "/my-trips", "Cancel Booking option visible"),
            ("Renter 1", "Click Cancel Booking", "/my-trips", "CancelBookingDialog: policy, refund estimate"),
            ("Renter 1", "Reason: Change of plans \u2192 Confirm", "/my-trips", "Toast: Cancellation requested"),
            ("Renter 1", "View booking in My Trips", "/my-trips", "Status: Cancelled, refund amount shown"),
            ("Owner 1", "Log in \u2192 check notification", "Any", "Notification: Booking cancelled"),
            ("Owner 1", "/owner-dashboard?tab=bookings-earnings", "/owner-dashboard", "Status: Cancelled, listing re-available"),
            ("Admin", "/admin?tab=bookings", "/admin", "Cancellation with refund details"),
            ("Admin", "Verify refund processed", "/admin", "Refund matches policy (flex/mod/strict)"),
        ]
    },
    {
        "id": "S-10", "name": "Owner Cancellation Flow",
        "tier": "Communication", "est_time": "10 min",
        "roles": "Owner 1, Renter 1, RAV Admin",
        "capability": "Owner cancels \u2192 Full refund \u2192 Penalty",
        "preconditions": "Confirmed booking between Owner 1 and Renter 1",
        "steps": [
            ("Owner 1", "/owner-dashboard?tab=bookings-earnings \u2192 select booking", "/owner-dashboard", "Cancel option visible"),
            ("Owner 1", "Click Cancel \u2192 dialog warns about penalty", "/owner-dashboard", "Warning: full refund + cancellation count"),
            ("Owner 1", "Select reason \u2192 Confirm", "/owner-dashboard", "Cancellation processed"),
            ("Renter 1", "/my-trips?tab=bookings", "/my-trips", "Status: Cancelled by Owner, full refund"),
            ("Renter 1", "Check notification", "Any", "Notification: Owner cancelled, full refund"),
            ("Admin", "/admin?tab=bookings", "/admin", "Owner cancellation logged, refund = 100%"),
            ("Owner 1", "Check dashboard", "/owner-dashboard", "Cancellation count incremented"),
        ]
    },
    {
        "id": "S-11", "name": "Dispute Resolution \u2014 Check-In Issue",
        "tier": "Communication", "est_time": "12 min",
        "roles": "Renter 1, RAV Admin",
        "capability": "Report issue \u2192 Evidence \u2192 Admin resolves",
        "preconditions": "Confirmed booking at/past check-in date",
        "steps": [
            ("Renter 1", "/checkin or /my-trips \u2192 select booking", "Various", "Report Issue button visible"),
            ("Renter 1", "Click Report Issue \u2192 ReportIssueDialog", "\u2014", "Category dropdown, description, evidence upload"),
            ("Renter 1", "Category: Room not as described, add description + photo \u2192 Submit", "\u2014", "Toast: Issue reported"),
            ("Renter 1", "/my-trips \u2192 booking detail", "/my-trips", "Dispute status: Under Review"),
            ("Admin", "/admin?tab=disputes", "/admin", "New dispute with category, description"),
            ("Admin", "Click dispute \u2192 view evidence thumbnails", "/admin", "Photos render correctly"),
            ("Admin", "Add notes, select resolution (full/partial refund) \u2192 Resolve", "/admin", "Status: Resolved"),
            ("Renter 1", "Check /my-trips + notification", "/my-trips", "Dispute resolved, refund details shown"),
        ]
    },
    {
        "id": "S-12", "name": "Owner Earnings & Payout Cycle",
        "tier": "Business Ops", "est_time": "10 min",
        "roles": "Owner 2 (Maria, Business), RAV Admin",
        "capability": "Completed booking \u2192 Escrow release \u2192 Payout",
        "preconditions": "Completed booking for Owner 2 with escrow released",
        "steps": [
            ("Owner 2", "Log in \u2192 /owner-dashboard (Dashboard tab)", "/owner-dashboard", "Earnings summary: total, pending, paid"),
            ("Owner 2", "Bookings & Earnings tab", "/owner-dashboard", "Completed bookings with amounts"),
            ("Owner 2", "Verify commission: total - 10% (Business tier)", "/owner-dashboard", "Commission calc correct"),
            ("Owner 2", "Check Stripe Connect status (StripeConnectBanner)", "/owner-dashboard", "Connect or Payouts Enabled state"),
            ("Admin", "/admin?tab=payouts", "/admin", "Owner 2 pending payout visible"),
            ("Admin", "Pay via Stripe \u2192 Confirm in AlertDialog", "/admin", "Payout: Processing/Paid"),
            ("Owner 2", "Refresh Bookings & Earnings", "/owner-dashboard", "Payout status: Paid"),
        ]
    },
    {
        "id": "S-13", "name": "Membership Upgrade \u2014 Owner",
        "tier": "Business Ops", "est_time": "8 min",
        "roles": "Owner 3 (James, Free \u2192 Pro)",
        "capability": "Free \u2192 Pro \u2192 Listing limit increase",
        "preconditions": "Owner 3 has Free tier (3 listing limit)",
        "steps": [
            ("Owner 3", "Log in \u2192 /owner-dashboard?tab=account", "/owner-dashboard", "Tier: Free, limit: 3"),
            ("Owner 3", "Click Upgrade Plan \u2192 MembershipPlans", "/owner-dashboard", "Tier comparison table"),
            ("Owner 3", "Select Pro tier \u2192 Stripe Checkout", "Stripe", "Pro plan at $10/month"),
            ("Owner 3", "Complete payment", "/subscription-success", "Subscription confirmed"),
            ("Owner 3", "Return to /owner-dashboard?tab=account", "/owner-dashboard", "Tier: Pro, limit: 10"),
            ("Owner 3", "Navigate to /list-property", "/list-property", "Can create new listing (was at limit)"),
            ("Owner 3", "Check My Listings", "/owner-dashboard", "Limit shows updated (e.g., 2/10)"),
        ]
    },
    {
        "id": "S-14", "name": "Membership Upgrade \u2014 Renter",
        "tier": "Business Ops", "est_time": "8 min",
        "roles": "Renter 4 (Free \u2192 Plus)",
        "capability": "Free \u2192 Plus \u2192 Perks visible",
        "preconditions": "Renter 4 has Free tier",
        "steps": [
            ("Renter 4", "Log in \u2192 /my-trips", "/my-trips", "Current tier shown, upgrade prompt"),
            ("Renter 4", "Click upgrade \u2192 MembershipPlans", "/my-trips", "Free vs Plus ($5) vs Premium ($15)"),
            ("Renter 4", "Select Plus \u2192 Stripe Checkout", "Stripe", "Plus at $5/month"),
            ("Renter 4", "Complete payment", "/subscription-success", "Subscription confirmed"),
            ("Renter 4", "Return to /my-trips", "/my-trips", "Tier: Plus, perks visible"),
        ]
    },
    {
        "id": "S-15", "name": "Referral Program End-to-End",
        "tier": "Business Ops", "est_time": "10 min",
        "roles": "Owner 1, New User (incognito), RAV Admin",
        "capability": "Generate code \u2192 Signup with ref \u2192 Track",
        "preconditions": "Owner 1 has referral code",
        "steps": [
            ("Owner 1", "/owner-dashboard?tab=account \u2192 Referral section", "/owner-dashboard", "ReferralDashboard with code + share link"),
            ("Owner 1", "Copy referral link", "/owner-dashboard", "Link copied"),
            ("New User", "Incognito \u2192 paste referral link \u2192 /signup?ref=REFxxx", "/signup", "Signup form with ref param captured"),
            ("New User", "Complete signup", "/signup", "Account created, referral recorded"),
            ("Admin", "Approve new user", "/admin", "User approved"),
            ("Owner 1", "Check ReferralDashboard", "/owner-dashboard", "Referral count incremented"),
        ]
    },
    {
        "id": "S-16", "name": "Search, Filter & Discovery",
        "tier": "Business Ops", "est_time": "10 min",
        "roles": "Renter 1 (Plus)",
        "capability": "Filters \u2192 Sort \u2192 Compare \u2192 Save search",
        "preconditions": "Active listings in seed data",
        "steps": [
            ("Renter 1", "Log in \u2192 /rentals", "/rentals", "Listing grid with filter bar"),
            ("Renter 1", "Filter: Orlando, 2+ bedrooms, $100-300/night", "/rentals", "Results narrow, count updates"),
            ("Renter 1", "Sort: Price Low to High", "/rentals", "Order correct"),
            ("Renter 1", "Sort: Price High to Low", "/rentals", "Order reverses"),
            ("Renter 1", "Click Save Search", "/rentals", "Toast: Search saved"),
            ("Renter 1", "Enable Compare mode", "/rentals", "Checkboxes appear on cards"),
            ("Renter 1", "Select 2-3 listings \u2192 Compare", "/rentals", "CompareListingsDialog with Best Value badges"),
            ("Renter 1", "Close \u2192 /my-trips?tab=favorites", "/my-trips", "Saved search visible"),
            ("Renter 1", "/destinations", "/destinations", "10 destinations with city counts"),
            ("Renter 1", "Click destination \u2192 city \u2192 listings", "/destinations/:slug/:city", "Filtered listings"),
        ]
    },
    {
        "id": "S-17", "name": "Notification Delivery & Preferences",
        "tier": "Communication", "est_time": "10 min",
        "roles": "Renter 1, Owner 1",
        "capability": "Trigger event \u2192 Notification arrives \u2192 Preferences",
        "preconditions": "Active interactions between Renter 1 and Owner 1",
        "steps": [
            ("Renter 1", "Click notification bell", "Any", "Unread count + dropdown list"),
            ("Renter 1", "Click View All", "/notifications", "Full feed with category filters"),
            ("Renter 1", "Filter by Bookings", "/notifications", "Only booking notifications"),
            ("Renter 1", "/settings/notifications", "/settings/notifications", "Per-type per-channel toggles"),
            ("Renter 1", "Toggle OFF email for Booking updates", "/settings/notifications", "Preference saved"),
            ("\u2014", "Trigger a booking event", "\u2014", "\u2014"),
            ("Renter 1", "Verify: in-app notification \u2705, email \u274c", "Various", "Channel preference respected"),
            ("Owner 1", "/settings/notifications", "/settings/notifications", "Owner-specific notification types"),
        ]
    },
    {
        "id": "S-18", "name": "Admin \u2014 Full Operations Cycle",
        "tier": "Admin & Platform", "est_time": "15 min",
        "roles": "RAV Admin, RAV Staff",
        "capability": "All admin tabs + staff boundary check",
        "preconditions": "Seed data loaded",
        "steps": [
            ("Admin", "Log in \u2192 /admin", "/admin?tab=overview", "Overview: KPIs, charts, pending counts"),
            ("Admin", "Click through Properties tab", "/admin?tab=properties", "Property list + edit button"),
            ("Admin", "Listings tab", "/admin?tab=listings", "Approve/reject queue"),
            ("Admin", "Bookings tab", "/admin?tab=bookings", "Booking list + filters"),
            ("Admin", "Escrow tab", "/admin?tab=escrow", "Verify/release controls"),
            ("Admin", "Issues tab", "/admin?tab=issues", "Check-in issues list"),
            ("Admin", "Disputes tab", "/admin?tab=disputes", "Dispute queue + evidence"),
            ("Admin", "Verifications tab", "/admin?tab=verifications", "Doc review queue"),
            ("Admin", "Users tab", "/admin?tab=users", "User search + roles"),
            ("Admin", "Approvals tab", "/admin?tab=pending-approvals", "Pending users + upgrades"),
            ("Admin", "Financials tab", "/admin?tab=financials", "Revenue, commissions"),
            ("Admin", "Tax tab", "/admin?tab=tax", "1099-K tracking"),
            ("Admin", "Payouts tab", "/admin?tab=payouts", "Payout queue"),
            ("Admin", "Memberships tab", "/admin?tab=memberships", "MRR, tier distribution"),
            ("Admin", "Settings tab", "/admin?tab=settings", "Commission, Staff-Only Mode"),
            ("Admin", "Voice tab", "/admin?tab=voice", "Quotas, usage charts"),
            ("Admin", "Resorts tab", "/admin?tab=resorts", "117 resorts, import tool"),
            ("Admin", "Notifications tab", "/admin?tab=notifications", "Catalog, events, delivery log"),
            ("Admin", "API Keys tab", "/admin?tab=api-keys", "Key management"),
            ("Admin", "Dev Tools tab (DEV only)", "/admin?tab=dev-tools", "Seed manager"),
            ("Staff", "Log in as dev-staff \u2192 /admin", "/admin", "Overview loads"),
            ("Staff", "Verify 10 operational tabs accessible", "/admin", "All 10 load correctly"),
            ("Staff", "Try /admin?tab=financials (direct URL)", "/admin", "Access denied or redirected"),
            ("Staff", "Verify admin-only tabs NOT in navigation", "/admin", "Financials/Tax/etc not visible"),
        ]
    },
    {
        "id": "S-19", "name": "Stripe Payment Edge Cases",
        "tier": "Core Marketplace", "est_time": "10 min",
        "roles": "Renter 1",
        "capability": "Decline, 3DS, expired card handling",
        "preconditions": "Active listing for checkout",
        "steps": [
            ("Renter 1", "Start checkout for a listing", "/checkout", "Payment form loads"),
            ("Renter 1", "Card: 4000 0000 0000 9995 (insufficient funds)", "/checkout", "Error: Insufficient funds"),
            ("Renter 1", "Card: 4000 0000 0000 0002 (decline)", "/checkout", "Error: Card declined"),
            ("Renter 1", "Card: 4000 0000 0000 0069 (expired)", "/checkout", "Error: Card expired"),
            ("Renter 1", "Card: 4000 0027 6000 3184 (3DS) \u2192 complete auth", "/checkout", "3DS popup \u2192 payment succeeds"),
            ("Renter 1", "New checkout \u2192 Card: 4000 0084 0000 1629 (3DS fail)", "/checkout", "3DS popup \u2192 payment declined"),
        ]
    },
    {
        "id": "S-20", "name": "Public Pages & Unauthenticated Access",
        "tier": "Foundation", "est_time": "10 min",
        "roles": "Guest (logged out)",
        "capability": "All public routes + protected route guards",
        "preconditions": "Logged out / incognito browser",
        "steps": [
            ("Guest", "Navigate to /", "/", "Landing page: hero, features, properties"),
            ("Guest", "/how-it-works", "/how-it-works", "Page loads with sections"),
            ("Guest", "/destinations", "/destinations", "10 destinations with counts"),
            ("Guest", "/faq", "/faq", "FAQ accordion loads"),
            ("Guest", "/terms", "/terms", "Terms renders"),
            ("Guest", "/privacy", "/privacy", "Privacy renders"),
            ("Guest", "/contact", "/contact", "Contact form loads"),
            ("Guest", "/tools", "/tools", "5 tools listed"),
            ("Guest", "/calculator", "/calculator", "SmartEarn works without login"),
            ("Guest", "/developers", "/developers", "API docs / Swagger loads"),
            ("Guest", "/user-guide", "/user-guide", "User guide renders"),
            ("Guest", "/login", "/login", "Login form + Google OAuth"),
            ("Guest", "/signup", "/signup", "Signup form loads"),
            ("Guest", "Try /rentals (protected)", "/login", "Redirected to login"),
            ("Guest", "Try /admin (protected)", "/login", "Redirected to login"),
            ("Guest", "Try /my-trips (protected)", "/login", "Redirected to login"),
            ("Guest", "/nonexistent-page", "404", "Custom 404 page"),
        ]
    },
    {
        "id": "S-21", "name": "Owner Multi-Listing Management",
        "tier": "Edge Cases", "est_time": "10 min",
        "roles": "Owner 3 (Free, 3 listing limit)",
        "capability": "Multiple listings, tier limits, drafts",
        "preconditions": "Owner 3 has Free tier (3 listing limit)",
        "steps": [
            ("Owner 3", "/owner-dashboard?tab=my-listings", "/owner-dashboard", "Listing count shown (e.g., 2/3)"),
            ("Owner 3", "/list-property \u2192 start filling \u2192 navigate away", "/list-property", "Form auto-saves to localStorage"),
            ("Owner 3", "Return to /list-property", "/list-property", "Draft restored from localStorage"),
            ("Owner 3", "Complete and submit listing", "/list-property", "Submitted (now at 3/3 limit)"),
            ("Owner 3", "Try /list-property again", "/list-property", "Warning: Reached limit, upgrade link"),
            ("Owner 3", "Edit existing listing in My Listings", "/owner-dashboard", "Edit dialog: nightly rate, dates, policy"),
            ("Owner 3", "Verify listing statuses", "/owner-dashboard", "Each shows correct status"),
        ]
    },
    {
        "id": "S-22", "name": "RAV Smart Tools Suite",
        "tier": "Foundation", "est_time": "10 min",
        "roles": "Guest (any user or logged out)",
        "capability": "SmartEarn, SmartCompare, SmartMatch, SmartBudget",
        "preconditions": "None",
        "steps": [
            ("Guest", "/tools", "/tools", "Hub: 5 tools with descriptions"),
            ("Guest", "/calculator \u2192 Hilton, $2400/yr, 1 week \u2192 Calculate", "/calculator", "Breakeven analysis + bar chart"),
            ("Guest", "Toggle Yield Estimator mode", "/calculator", "ROI % + projected earnings"),
            ("Guest", "/tools/cost-comparator \u2192 hotel $300/night, 7 nights", "/tools/cost-comparator", "Hotel vs timeshare vs RAV comparison"),
            ("Guest", "/tools/resort-quiz \u2192 answer questions", "/tools/resort-quiz", "Results: matched resorts with scores"),
            ("Guest", "/tools/budget-planner \u2192 fill inputs", "/tools/budget-planner", "Budget breakdown rendered"),
            ("Guest", "Verify breadcrumbs navigate to /tools", "All tool pages", "Breadcrumbs work"),
        ]
    },
    {
        "id": "S-23", "name": "Date Proposal & Flexible Booking",
        "tier": "Core Marketplace", "est_time": "12 min",
        "roles": "Renter 1, Owner 1",
        "capability": "Propose alt dates \u2192 Owner reviews \u2192 Accept",
        "preconditions": "Active bidding listing from Owner 1",
        "steps": [
            ("Renter 1", "Find bidding listing \u2192 Make an Offer", "/property/:id", "BidFormDialog in bid mode"),
            ("Renter 1", "Switch to Propose Different Dates mode", "/property/:id", "Date pickers + auto-computed bid"),
            ("Renter 1", "Select different check-in/out, verify auto-calc", "/property/:id", "Bid = nightly_rate \u00d7 nights"),
            ("Renter 1", "Add message \u2192 Submit", "/property/:id", "Toast: Date proposal submitted"),
            ("Owner 1", "My Listings \u2192 BidsManagerDialog", "/owner-dashboard", "Date proposal with proposed dates + amount"),
            ("Owner 1", "Accept or counter the date proposal", "/owner-dashboard", "Response sent"),
            ("Renter 1", "/my-trips?tab=offers", "/my-trips", "Date proposal response visible"),
        ]
    },
    {
        "id": "S-24", "name": "Admin Property & Listing Editing",
        "tier": "Admin & Platform", "est_time": "10 min",
        "roles": "RAV Admin",
        "capability": "Edit property/listing details, audit trail",
        "preconditions": "Existing properties and listings in seed data",
        "steps": [
            ("Admin", "/admin?tab=properties \u2192 select property", "/admin", "Property details visible"),
            ("Admin", "Click Edit \u2192 AdminPropertyEditDialog", "/admin", "Fields: brand, resort, bedrooms, etc."),
            ("Admin", "Change bedrooms 2\u21923 \u2192 Save", "/admin", "Success toast, property updated"),
            ("Admin", "Verify audit trail: edited by + timestamp", "/admin", "last_edited_by/at populated"),
            ("Admin", "/admin?tab=listings \u2192 select active listing", "/admin", "Listing with pricing"),
            ("Admin", "Click Edit \u2192 AdminListingEditDialog", "/admin", "Live price calc visible"),
            ("Admin", "Change nightly rate \u2192 observe recalculation", "/admin", "Owner price, markup, final update live"),
            ("Admin", "Save changes", "/admin", "Listing updated"),
            ("Admin", "Try editing booked/completed listing", "/admin", "Edit disabled or blocked"),
        ]
    },
    {
        "id": "S-25", "name": "Executive Dashboard & Analytics",
        "tier": "Admin & Platform", "est_time": "8 min",
        "roles": "RAV Owner (dev-owner@)",
        "capability": "Strategic metrics, market data",
        "preconditions": "Seed data loaded",
        "steps": [
            ("RAV Owner", "/executive-dashboard", "/executive-dashboard", "Dashboard loads with data"),
            ("RAV Owner", "Verify KPI cards: GMV, listings, users, conversion", "/executive-dashboard", "Numbers render from seed data"),
            ("RAV Owner", "Check liquidity score", "/executive-dashboard", "Score + explanation renders"),
            ("RAV Owner", "Verify unit economics section", "/executive-dashboard", "Commission, avg booking, CAC"),
            ("RAV Owner", "Check market benchmarks", "/executive-dashboard", "Section renders"),
            ("Staff", "Log in as dev-staff \u2192 /executive-dashboard", "/executive-dashboard", "Access denied or redirected"),
        ]
    },
    {
        "id": "S-26", "name": "Voice Search & Quota Management",
        "tier": "Admin & Platform", "est_time": "10 min",
        "roles": "Renter 1 (Plus, 25/day), RAV Admin",
        "capability": "Voice search \u2192 Quota tracking \u2192 Admin override",
        "preconditions": "VAPI configured on DEV",
        "steps": [
            ("Renter 1", "/rentals \u2192 click voice search icon", "/rentals", "Voice UI activates"),
            ("Renter 1", "Search: 'Beach vacation in Orlando'", "/rentals", "Results filter, search logged"),
            ("Admin", "/admin?tab=voice \u2192 VoiceUsageDashboard", "/admin", "Renter 1's search in chart"),
            ("Admin", "VoiceUserOverrideManager \u2192 find Renter 1", "/admin", "User found"),
            ("Admin", "Set custom quota: 50/day", "/admin", "Override saved"),
            ("Admin", "Alternatively: disable voice for user", "/admin", "Disabled flag saved"),
            ("Renter 1", "If disabled: try voice search", "/rentals", "Voice search disabled message"),
            ("Admin", "VoiceObservability \u2192 log viewer", "/admin", "Recent logs with query, duration, count"),
        ]
    },
    {
        "id": "S-27", "name": "API Key & Developer Portal",
        "tier": "Admin & Platform", "est_time": "10 min",
        "roles": "RAV Admin",
        "capability": "Create key \u2192 Test endpoint \u2192 Revoke",
        "preconditions": "None",
        "steps": [
            ("Admin", "/admin?tab=api-keys", "/admin", "API key management loads"),
            ("Admin", "Create Key: name 'Test Key', tier 'partner', optional IPs", "/admin", "Key shown ONCE \u2014 copy it"),
            ("Admin", "Verify key in list (masked rav_...)", "/admin", "Key listed with date, tier"),
            ("Guest", "/developers", "/developers", "Swagger UI loads"),
            ("Admin", "Test API: GET /api/listings with key header", "\u2014", "Returns JSON, rate limit headers"),
            ("Admin", "Revoke key \u2192 confirm AlertDialog", "/admin", "Key revoked"),
            ("Admin", "Test same API call with revoked key", "\u2014", "401 Unauthorized"),
        ]
    },
    {
        "id": "S-28", "name": "iCal Export & Calendar Integration",
        "tier": "Admin & Platform", "est_time": "5 min",
        "roles": "Owner 1 (Alex)",
        "capability": "Export bookings \u2192 Verify ICS file",
        "preconditions": "Owner 1 has bookings",
        "steps": [
            ("Owner 1", "/owner-dashboard?tab=bookings-earnings", "/owner-dashboard", "Booking list visible"),
            ("Owner 1", "Click Export Calendar", "/owner-dashboard", "ICS file downloads"),
            ("Owner 1", "Open ICS in text editor", "Local", "Valid iCal: BEGIN:VCALENDAR, events"),
            ("Owner 1", "Verify dates match bookings, property names correct", "Local", "Data accurate"),
        ]
    },
    {
        "id": "S-29", "name": "GDPR Data Export & Deletion",
        "tier": "Edge Cases", "est_time": "8 min",
        "roles": "Renter 50 (expendable account)",
        "capability": "Request export \u2192 Deletion \u2192 Grace period",
        "preconditions": "Use renter050@ (can afford to lose)",
        "steps": [
            ("Renter 50", "Log in \u2192 /account", "/account", "Account settings loads"),
            ("Renter 50", "Click Export My Data", "/account", "Data export (JSON/download)"),
            ("Renter 50", "Click Delete My Account", "/account", "Warning: 14-day grace, irreversible"),
            ("Renter 50", "Confirm deletion", "/account", "Account marked for deletion, logged out"),
            ("Renter 50", "Try to log back in", "/login", "Grace period message or cancel option"),
        ]
    },
    {
        "id": "S-30", "name": "Complete Marketplace Cycle",
        "tier": "Master Validation", "est_time": "25 min",
        "roles": "Owner 1, Renter 1, RAV Admin",
        "capability": "Full lifecycle: list \u2192 bid \u2192 book \u2192 confirm \u2192 review \u2192 payout",
        "preconditions": "Reseeded DEV data recommended",
        "steps": [
            # Phase A
            ("Owner 1", "/list-property \u2192 Create listing: Hilton Las Vegas, 2BR, 7 nights, $250/night, Open for Bidding", "/list-property", "Submitted, Pending Approval"),
            # Phase B
            ("Admin", "/admin?tab=listings \u2192 Approve Owner 1's listing", "/admin", "Status: Active"),
            # Phase C
            ("Renter 1", "/rentals \u2192 search Las Vegas \u2192 find new listing", "/rentals", "Listing with bidding badge"),
            ("Renter 1", "Ask the Owner: 'Is parking included?'", "/property/:id", "Inquiry sent"),
            ("Owner 1", "Check notification \u2192 reply: 'Yes, free parking!'", "InquiryThread", "Reply sent"),
            ("Renter 1", "View reply \u2192 Place Bid at $1,500", "/property/:id", "Bid submitted"),
            # Phase D
            ("Owner 1", "BidsManagerDialog \u2192 counter at $1,750", "/owner-dashboard", "Counter sent"),
            ("Renter 1", "Accept counter in My Trips", "/my-trips", "Redirected to /checkout at $1,750"),
            ("Renter 1", "Pay with Stripe test card", "/booking-success", "Booking: Payment \u2705, Owner \u23f3"),
            # Phase E
            ("Owner 1", "Confirm booking \u2192 RST confirmation number", "/owner-dashboard", "Booking confirmed"),
            ("Renter 1", "/my-trips \u2192 Confirmed status", "/my-trips", "Timeline updated"),
            # Phase F
            ("Admin", "/admin?tab=escrow \u2192 verify \u2192 release", "/admin", "Escrow released"),
            # Phase G
            ("Renter 1", "Booking message: 'What time is check-in?'", "/my-trips", "Message sent"),
            ("Owner 1", "Reply: 'Check-in at 4pm'", "/owner-dashboard", "Realtime delivery"),
            # Phase H
            ("Renter 1", "Submit review: 5 stars, 'Amazing!'", "/my-trips", "Review submitted"),
            ("Owner 1", "Check earnings: $1,750 - 13% = $1,522.50", "/owner-dashboard", "Commission calc correct"),
            ("Admin", "/admin?tab=payouts \u2192 verify \u2192 initiate payout", "/admin", "Payout processed"),
            ("Owner 1", "Verify payout status: Paid", "/owner-dashboard", "Payout: Paid"),
        ]
    },
]

# Number of pre-built run columns
NUM_RUN_COLUMNS = 10


def create_workbook():
    wb = openpyxl.Workbook()

    # ── 1. DASHBOARD SHEET ──
    create_dashboard(wb)

    # ── 2. RUN LOG SHEET ──
    create_run_log(wb)

    # ── 3. SCENARIO SUMMARY SHEET ──
    create_scenario_summary(wb)

    # ── 4. INDIVIDUAL SCENARIO SHEETS (S-01 through S-30) ──
    for scenario in SCENARIOS:
        create_scenario_sheet(wb, scenario)

    # ── 5. REFERENCE SHEET ──
    create_reference_sheet(wb)

    # ── 6. COVERAGE MATRIX ──
    create_coverage_matrix(wb)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    return wb


def _kpi_box(ws, start_col_letter, end_col_letter, row, label, value, color, fmt='0'):
    """Render a single KPI card (label on top, big value below)."""
    start_col = openpyxl.utils.column_index_from_string(start_col_letter)
    for r in range(row, row + 3):
        for c_letter in [start_col_letter, end_col_letter]:
            c = openpyxl.utils.column_index_from_string(c_letter)
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid')
            cell.border = Border(
                left=Side(style='thin', color=GRAY_BORDER) if c_letter == start_col_letter else Side(style=None),
                right=Side(style='thin', color=GRAY_BORDER) if c_letter == end_col_letter else Side(style=None),
                top=Side(style='thin', color=GRAY_BORDER) if r == row else Side(style=None),
                bottom=Side(style='thin', color=GRAY_BORDER) if r == row + 2 else Side(style=None),
            )
    ws.merge_cells(f'{start_col_letter}{row}:{end_col_letter}{row}')
    ws.cell(row=row, column=start_col).value = label
    ws.cell(row=row, column=start_col).font = Font(name='Calibri', color=GRAY_TEXT, size=10)
    ws.cell(row=row, column=start_col).alignment = Alignment(horizontal='center')

    ws.merge_cells(f'{start_col_letter}{row+1}:{end_col_letter}{row+1}')
    ws.cell(row=row+1, column=start_col).value = value
    ws.cell(row=row+1, column=start_col).font = Font(name='Calibri', bold=True, color=color, size=28)
    ws.cell(row=row+1, column=start_col).alignment = Alignment(horizontal='center')
    ws.cell(row=row+1, column=start_col).number_format = fmt


def create_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = TEAL

    # ── Page setup — wider grid for charts ──
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 3

    LAST_COL = 10  # J
    LAST_COL_LETTER = 'J'

    # The "Latest Status" column in Summary is column R (8 + NUM_RUN_COLUMNS = 18)
    latest_col_letter = get_column_letter(8 + NUM_RUN_COLUMNS)  # R
    # Totals row in Summary is row 33 (30 scenarios + 2 header rows + 1)
    totals_row_summary = len(SCENARIOS) + 3  # 33

    # ── Header Banner ──
    for col in range(1, LAST_COL + 2):
        ws.cell(row=1, column=col).fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type='solid')
        ws.cell(row=2, column=col).fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type='solid')
        ws.cell(row=3, column=col).fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type='solid')

    ws.merge_cells(f'B1:{LAST_COL_LETTER}1')
    ws.cell(row=1, column=2).value = "RENT-A-VACATION"
    ws.cell(row=1, column=2).font = Font(name='Calibri', bold=True, color=WHITE, size=22)
    ws.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f'B2:{LAST_COL_LETTER}2')
    ws.cell(row=2, column=2).value = "QA Scenario Test Dashboard"
    ws.cell(row=2, column=2).font = Font(name='Calibri', color=TEAL_MED, size=14)
    ws.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f'B3:{LAST_COL_LETTER}3')
    ws.cell(row=3, column=2).value = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    ws.cell(row=3, column=2).font = Font(name='Calibri', color=TEAL_MED, size=10)
    ws.cell(row=3, column=2).alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18

    # ══════════════════════════════════════════
    # ROW 1 KPIs (row 5-9): 5 cards across
    # ══════════════════════════════════════════
    ws.merge_cells('B5:C5')
    ws.cell(row=5, column=2).value = "LATEST RUN SUMMARY"
    ws.cell(row=5, column=2).font = section_font
    ws.cell(row=5, column=2).alignment = Alignment(horizontal='left')

    row = 7
    _kpi_box(ws, 'B', 'C', row, "Total Scenarios",
             '=COUNTIF(Summary!A3:A200,"S-*")', TEAL)
    _kpi_box(ws, 'D', 'E', row, "Passed",
             f'=Summary!B{totals_row_summary}', GREEN)
    _kpi_box(ws, 'F', 'G', row, "Failed",
             f'=Summary!C{totals_row_summary}', RED)
    _kpi_box(ws, 'H', 'I', row, "Blocked",
             f'=Summary!D{totals_row_summary}', AMBER)

    # ── ROW 2 KPIs (row 11-13) ──
    ws.row_dimensions[10].height = 8  # spacer
    row = 11
    _kpi_box(ws, 'B', 'C', row, "Skipped",
             f'=Summary!E{totals_row_summary}', GRAY_TEXT)
    _kpi_box(ws, 'D', 'E', row, "Pass Rate",
             f'=IF(Summary!B{totals_row_summary}+Summary!C{totals_row_summary}>0,'
             f'Summary!B{totals_row_summary}/(Summary!B{totals_row_summary}+Summary!C{totals_row_summary}),0)',
             TEAL, '0%')
    _kpi_box(ws, 'F', 'G', row, "Execution Coverage",
             f'=IF(COUNTIF(Summary!A3:A200,"S-*")>0,'
             f'COUNTIF(Summary!{latest_col_letter}3:{latest_col_letter}200,"<>")/COUNTIF(Summary!A3:A200,"S-*"),0)',
             BLUE, '0%')
    _kpi_box(ws, 'H', 'I', row, "Total Steps",
             f'=SUM(Summary!F3:F{totals_row_summary - 1})', DARK)

    # ══════════════════════════════════════════
    # CHART DATA AREA (hidden, rows 40-55 — used by charts)
    # ══════════════════════════════════════════
    # We place chart source data below the visible area so charts
    # reference clean labeled ranges.

    cd_row = 50  # chart data start row

    # ── Pie chart data: Status Distribution ──
    ws.cell(row=cd_row, column=2).value = "Status"
    ws.cell(row=cd_row, column=3).value = "Count"
    ws.cell(row=cd_row, column=2).font = Font(name='Calibri', bold=True, size=9, color=GRAY_TEXT)
    ws.cell(row=cd_row, column=3).font = Font(name='Calibri', bold=True, size=9, color=GRAY_TEXT)

    pie_labels = ["Passed", "Failed", "Blocked", "Skipped", "Not Run"]
    pie_formulas = [
        f'=Summary!B{totals_row_summary}',
        f'=Summary!C{totals_row_summary}',
        f'=Summary!D{totals_row_summary}',
        f'=Summary!E{totals_row_summary}',
        f'=COUNTIF(Summary!A3:A200,"S-*")-Summary!B{totals_row_summary}-Summary!C{totals_row_summary}'
        f'-Summary!D{totals_row_summary}-Summary!E{totals_row_summary}',
    ]
    for i, (lbl, fml) in enumerate(zip(pie_labels, pie_formulas)):
        ws.cell(row=cd_row + 1 + i, column=2).value = lbl
        ws.cell(row=cd_row + 1 + i, column=2).font = Font(name='Calibri', size=9, color=GRAY_TEXT)
        ws.cell(row=cd_row + 1 + i, column=3).value = fml
        ws.cell(row=cd_row + 1 + i, column=3).font = Font(name='Calibri', size=9, color=GRAY_TEXT)

    # ── Tier bar chart data ──
    tier_cd_row = cd_row + 8
    ws.cell(row=tier_cd_row, column=2).value = "Tier"
    ws.cell(row=tier_cd_row, column=3).value = "Passed"
    ws.cell(row=tier_cd_row, column=4).value = "Failed"
    ws.cell(row=tier_cd_row, column=5).value = "Blocked"
    for c in range(2, 6):
        ws.cell(row=tier_cd_row, column=c).font = Font(name='Calibri', bold=True, size=9, color=GRAY_TEXT)

    # ── Run history trend data ──
    run_cd_row = cd_row + 18
    ws.cell(row=run_cd_row, column=2).value = "Run"
    ws.cell(row=run_cd_row, column=3).value = "Passed"
    ws.cell(row=run_cd_row, column=4).value = "Failed"
    ws.cell(row=run_cd_row, column=5).value = "Blocked"
    for c in range(2, 6):
        ws.cell(row=run_cd_row, column=c).font = Font(name='Calibri', bold=True, size=9, color=GRAY_TEXT)

    # Run history: count PASS/FAIL/BLOCKED per run column in Summary
    for run_i in range(NUM_RUN_COLUMNS):
        r = run_cd_row + 1 + run_i
        run_col_letter = get_column_letter(8 + run_i)  # H, I, J, ...
        last_scenario_row = len(SCENARIOS) + 2
        ws.cell(row=r, column=2).value = f"Run {run_i + 1}"
        ws.cell(row=r, column=2).font = Font(name='Calibri', size=9, color=GRAY_TEXT)
        ws.cell(row=r, column=3).value = f'=COUNTIF(Summary!{run_col_letter}3:{run_col_letter}{last_scenario_row},"PASS")'
        ws.cell(row=r, column=3).font = Font(name='Calibri', size=9, color=GRAY_TEXT)
        ws.cell(row=r, column=4).value = f'=COUNTIF(Summary!{run_col_letter}3:{run_col_letter}{last_scenario_row},"FAIL")'
        ws.cell(row=r, column=4).font = Font(name='Calibri', size=9, color=GRAY_TEXT)
        ws.cell(row=r, column=5).value = f'=COUNTIF(Summary!{run_col_letter}3:{run_col_letter}{last_scenario_row},"BLOCKED")'
        ws.cell(row=r, column=5).font = Font(name='Calibri', size=9, color=GRAY_TEXT)

    # ══════════════════════════════════════════
    # CHARTS SECTION (row 15)
    # ══════════════════════════════════════════
    row = 15
    ws.merge_cells(f'B{row}:{LAST_COL_LETTER}{row}')
    ws.cell(row=row, column=2).value = "STATUS DISTRIBUTION & TREND"
    ws.cell(row=row, column=2).font = section_font

    # ── Pie Chart: Status Distribution ──
    pie = PieChart()
    pie.title = "Status Distribution"
    pie.style = 10
    pie.width = 14
    pie.height = 12
    cats = Reference(ws, min_col=2, min_row=cd_row + 1, max_row=cd_row + 5)
    vals = Reference(ws, min_col=3, min_row=cd_row, max_row=cd_row + 5)
    pie.add_data(vals, titles_from_data=True)
    pie.set_categories(cats)
    # Brand colors: green, red, amber, gray, light gray
    colors = [GREEN, RED, AMBER, GRAY_TEXT, GRAY_BORDER]
    for i, color in enumerate(colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)
    pie.legend.position = 'b'
    ws.add_chart(pie, "B16")

    # ── Stacked Bar Chart: Run History Trend ──
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "stacked"
    bar.title = "Run History Trend"
    bar.style = 10
    bar.width = 18
    bar.height = 12
    bar.y_axis.title = "Scenarios"
    bar.x_axis.title = "Test Run"

    cats = Reference(ws, min_col=2, min_row=run_cd_row + 1, max_row=run_cd_row + NUM_RUN_COLUMNS)
    bar.set_categories(cats)

    for col_idx, (label, color) in enumerate(
        [("Passed", GREEN), ("Failed", RED), ("Blocked", AMBER)], start=3
    ):
        vals = Reference(ws, min_col=col_idx, min_row=run_cd_row, max_row=run_cd_row + NUM_RUN_COLUMNS)
        bar.add_data(vals, titles_from_data=True)
        bar.series[-1].graphicalProperties.solidFill = color

    ws.add_chart(bar, "F16")

    # ══════════════════════════════════════════
    # TIER BREAKDOWN TABLE (row 32)
    # ══════════════════════════════════════════
    row = 32
    ws.merge_cells(f'B{row}:{LAST_COL_LETTER}{row}')
    ws.cell(row=row, column=2).value = "RESULTS BY TIER"
    ws.cell(row=row, column=2).font = section_font

    row = 34
    tier_headers = ["Tier", "Scenarios", "Passed", "Failed", "Blocked", "Pass Rate", "Progress"]
    for i, h in enumerate(tier_headers):
        col = i + 2
        cell = ws.cell(row=row, column=col)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    tiers = [
        ("Foundation", [1, 3, 20, 22]),
        ("Core Marketplace", [4, 5, 6, 19, 23]),
        ("Communication", [7, 8, 9, 10, 11, 17]),
        ("Business Ops", [12, 13, 14, 15, 16]),
        ("Admin & Platform", [2, 18, 24, 25, 26, 27, 28]),
        ("Edge Cases", [21, 29]),
        ("Master Validation", [30]),
    ]

    for t_idx, (tier_name, scenario_rows) in enumerate(tiers):
        r = row + 1 + t_idx
        ws.cell(row=r, column=2).value = tier_name
        ws.cell(row=r, column=2).font = Font(name='Calibri', bold=True, color=DARK, size=11)
        ws.cell(row=r, column=2).border = thin_border

        ws.cell(row=r, column=3).value = len(scenario_rows)
        ws.cell(row=r, column=3).alignment = center_align
        ws.cell(row=r, column=3).border = thin_border

        pass_parts, fail_parts, blocked_parts = [], [], []
        for sr in scenario_rows:
            summary_row = sr + 2
            pass_parts.append(f'IF(Summary!{latest_col_letter}{summary_row}="PASS",1,0)')
            fail_parts.append(f'IF(Summary!{latest_col_letter}{summary_row}="FAIL",1,0)')
            blocked_parts.append(f'IF(Summary!{latest_col_letter}{summary_row}="BLOCKED",1,0)')

        ws.cell(row=r, column=4).value = f'={"+".join(pass_parts)}'
        ws.cell(row=r, column=4).alignment = center_align
        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=4).font = Font(name='Calibri', color=GREEN)

        ws.cell(row=r, column=5).value = f'={"+".join(fail_parts)}'
        ws.cell(row=r, column=5).alignment = center_align
        ws.cell(row=r, column=5).border = thin_border
        ws.cell(row=r, column=5).font = Font(name='Calibri', color=RED)

        ws.cell(row=r, column=6).value = f'={"+".join(blocked_parts)}'
        ws.cell(row=r, column=6).alignment = center_align
        ws.cell(row=r, column=6).border = thin_border
        ws.cell(row=r, column=6).font = Font(name='Calibri', color=AMBER)

        # Pass Rate
        ws.cell(row=r, column=7).value = (
            f'=IF(D{r}+E{r}>0,D{r}/(D{r}+E{r}),0)'
        )
        ws.cell(row=r, column=7).number_format = '0%'
        ws.cell(row=r, column=7).alignment = center_align
        ws.cell(row=r, column=7).border = thin_border
        ws.cell(row=r, column=7).font = Font(name='Calibri', bold=True, color=TEAL)

        # Progress (pass rate as visual indicator)
        ws.cell(row=r, column=8).value = f'=G{r}'
        ws.cell(row=r, column=8).number_format = '0%'
        ws.cell(row=r, column=8).alignment = center_align
        ws.cell(row=r, column=8).border = thin_border

        # Also populate tier bar chart data
        tier_data_r = tier_cd_row + 1 + t_idx
        ws.cell(row=tier_data_r, column=2).value = tier_name
        ws.cell(row=tier_data_r, column=2).font = Font(name='Calibri', size=9, color=GRAY_TEXT)
        ws.cell(row=tier_data_r, column=3).value = f'=D{r}'
        ws.cell(row=tier_data_r, column=3).font = Font(name='Calibri', size=9, color=GRAY_TEXT)
        ws.cell(row=tier_data_r, column=4).value = f'=E{r}'
        ws.cell(row=tier_data_r, column=4).font = Font(name='Calibri', size=9, color=GRAY_TEXT)
        ws.cell(row=tier_data_r, column=5).value = f'=F{r}'
        ws.cell(row=tier_data_r, column=5).font = Font(name='Calibri', size=9, color=GRAY_TEXT)

    # Data bars on the Progress column
    tier_first_row = row + 1
    tier_last_row = row + len(tiers)
    ws.conditional_formatting.add(
        f'H{tier_first_row}:H{tier_last_row}',
        DataBarRule(
            start_type='num', start_value=0,
            end_type='num', end_value=1,
            color=TEAL,
            showValue=True,
        )
    )

    # Alternating rows for tier table
    for t_idx in range(len(tiers)):
        r = row + 1 + t_idx
        if t_idx % 2 == 0:
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = PatternFill(
                    start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid'
                )

    # ── Tier Bar Chart ──
    tier_bar = BarChart()
    tier_bar.type = "col"
    tier_bar.grouping = "stacked"
    tier_bar.title = "Results by Tier"
    tier_bar.style = 10
    tier_bar.width = 18
    tier_bar.height = 12
    tier_bar.y_axis.title = "Scenarios"

    cats = Reference(ws, min_col=2, min_row=tier_cd_row + 1, max_row=tier_cd_row + len(tiers))
    tier_bar.set_categories(cats)
    for col_idx, (label, color) in enumerate(
        [("Passed", GREEN), ("Failed", RED), ("Blocked", AMBER)], start=3
    ):
        vals = Reference(ws, min_col=col_idx, min_row=tier_cd_row, max_row=tier_cd_row + len(tiers))
        tier_bar.add_data(vals, titles_from_data=True)
        tier_bar.series[-1].graphicalProperties.solidFill = color

    ws.add_chart(tier_bar, "B43")

    # ══════════════════════════════════════════
    # SCENARIO HEALTH HEATMAP (row 43, right side)
    # ══════════════════════════════════════════
    # We show a compact "scenario scorecard" — S-01 through S-30 in a grid
    # with their latest status color-coded. This gives a quick visual map.

    hm_row = 60
    ws.merge_cells(f'B{hm_row}:{LAST_COL_LETTER}{hm_row}')
    ws.cell(row=hm_row, column=2).value = "SCENARIO HEALTH MAP"
    ws.cell(row=hm_row, column=2).font = section_font

    # 6 columns x 5 rows = 30 scenarios
    hm_cols = 6
    hm_start_row = hm_row + 1
    for i, s in enumerate(SCENARIOS):
        grid_row = hm_start_row + (i // hm_cols)
        grid_col = 2 + (i % hm_cols)
        summary_data_row = i + 3  # row 3 = S-01

        cell = ws.cell(row=grid_row, column=grid_col)
        cell.value = s["id"]
        cell.font = Font(name='Calibri', bold=True, size=10, color=DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color=GRAY_BORDER),
            right=Side(style='thin', color=GRAY_BORDER),
            top=Side(style='thin', color=GRAY_BORDER),
            bottom=Side(style='thin', color=GRAY_BORDER),
        )
        ws.row_dimensions[grid_row].height = 28

    # Conditional formatting on heatmap cells based on Summary latest status
    # We use helper cells (row 70+) that mirror Summary latest status
    helper_row_start = 75
    for i, s in enumerate(SCENARIOS):
        grid_row = hm_start_row + (i // hm_cols)
        grid_col = 2 + (i % hm_cols)
        summary_data_row = i + 3

        # Store the status lookup in a hidden helper row
        helper_r = helper_row_start + i
        ws.cell(row=helper_r, column=2).value = f'=Summary!{latest_col_letter}{summary_data_row}'
        ws.cell(row=helper_r, column=2).font = Font(name='Calibri', size=1, color=WHITE)

        # Instead of complex indirect formatting, we make each cell
        # show the scenario ID and we'll apply formatting per-cell
        # using a formula-based approach in the cell itself
        cell_ref = f'{get_column_letter(grid_col)}{grid_row}'

    # Apply conditional formatting to the entire heatmap grid area
    hm_end_row = hm_start_row + ((len(SCENARIOS) - 1) // hm_cols)
    hm_end_col = 2 + hm_cols - 1

    # We need the heatmap to color based on the MATCHING Summary row
    # Since openpyxl CF can't do cross-sheet INDIRECT easily, we use
    # a practical approach: place a "status mirror" in column J (hidden)
    # and format based on that.
    mirror_col = LAST_COL  # column J
    for i in range(len(SCENARIOS)):
        grid_row = hm_start_row + (i // hm_cols)
        summary_data_row = i + 3
        # We'll use a different approach: put status in the cell value
        # via formula and use CF on the value

    # Simpler approach: make each heatmap cell SHOW the status text
    # and color it. This is more useful than just showing "S-01".
    for i, s in enumerate(SCENARIOS):
        grid_row = hm_start_row + (i // hm_cols)
        grid_col = 2 + (i % hm_cols)
        summary_data_row = i + 3

        cell = ws.cell(row=grid_row, column=grid_col)
        # Show "S-01" with status below via formula
        cell.value = f'=IF(Summary!{latest_col_letter}{summary_data_row}<>"","{s["id"]}" & CHAR(10) & Summary!{latest_col_letter}{summary_data_row},"{s["id"]}")'
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Apply CF to heatmap range (check if cell CONTAINS "PASS", "FAIL", etc.)
    hm_range = f'B{hm_start_row}:{get_column_letter(2 + hm_cols - 1)}{hm_end_row}'

    ws.conditional_formatting.add(hm_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("PASS",B{hm_start_row}))'],
        fill=PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type='solid'),
        font=Font(color=GREEN, bold=True, size=10)
    ))
    ws.conditional_formatting.add(hm_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("FAIL",B{hm_start_row}))'],
        fill=PatternFill(start_color=RED_LIGHT, end_color=RED_LIGHT, fill_type='solid'),
        font=Font(color=RED, bold=True, size=10)
    ))
    ws.conditional_formatting.add(hm_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("BLOCKED",B{hm_start_row}))'],
        fill=PatternFill(start_color=AMBER_LIGHT, end_color=AMBER_LIGHT, fill_type='solid'),
        font=Font(color=AMBER, bold=True, size=10)
    ))
    ws.conditional_formatting.add(hm_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("SKIP",B{hm_start_row}))'],
        fill=PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid'),
        font=Font(color=GRAY_TEXT, size=10)
    ))

    # ══════════════════════════════════════════
    # DEFECT SUMMARY (row after heatmap)
    # ══════════════════════════════════════════
    defect_row = hm_end_row + 2
    ws.merge_cells(f'B{defect_row}:{LAST_COL_LETTER}{defect_row}')
    ws.cell(row=defect_row, column=2).value = "DEFECT TRACKER"
    ws.cell(row=defect_row, column=2).font = section_font

    defect_hdr_row = defect_row + 1
    defect_headers = ["#", "Scenario", "Step", "Severity", "GitHub Issue", "Status", "Found Date", "Resolved Date", "Notes"]
    defect_widths_map = {1: 5, 2: 10, 3: 6, 4: 12, 5: 16, 6: 12, 7: 14, 8: 14, 9: 30}
    for i, h in enumerate(defect_headers):
        ws.cell(row=defect_hdr_row, column=i + 2).value = h
    style_header_row(ws, defect_hdr_row, len(defect_headers) + 1)
    # Shift headers to start at col B
    for i, h in enumerate(defect_headers):
        ws.cell(row=defect_hdr_row, column=i + 2).value = h

    # Pre-fill 15 defect rows
    for i in range(15):
        r = defect_hdr_row + 1 + i
        ws.cell(row=r, column=2).value = i + 1
        ws.cell(row=r, column=2).alignment = center_align
        ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=2).font = body_font
        for c in range(3, 11):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = body_font
            cell.alignment = center_align if c < 10 else body_align
        if i % 2 == 0:
            for c in range(2, 11):
                ws.cell(row=r, column=c).fill = PatternFill(
                    start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid'
                )

    # Severity dropdown
    from openpyxl.worksheet.datavalidation import DataValidation
    sev_dv = DataValidation(type="list", formula1='"Critical,Major,Minor,Cosmetic"', showDropDown=False)
    ws.add_data_validation(sev_dv)
    sev_dv.add(f'E{defect_hdr_row+1}:E{defect_hdr_row+15}')

    # Defect status dropdown
    def_status_dv = DataValidation(type="list", formula1='"Open,In Progress,Fixed,Verified,Won\'t Fix"', showDropDown=False)
    ws.add_data_validation(def_status_dv)
    def_status_dv.add(f'G{defect_hdr_row+1}:G{defect_hdr_row+15}')

    # Conditional formatting for severity
    ws.conditional_formatting.add(f'E{defect_hdr_row+1}:E{defect_hdr_row+15}', CellIsRule(
        operator='equal', formula=['"Critical"'], fill=fail_fill,
        font=Font(color=RED, bold=True)))
    ws.conditional_formatting.add(f'E{defect_hdr_row+1}:E{defect_hdr_row+15}', CellIsRule(
        operator='equal', formula=['"Major"'], fill=blocked_fill,
        font=Font(color=AMBER, bold=True)))

    # Defect status formatting
    ws.conditional_formatting.add(f'G{defect_hdr_row+1}:G{defect_hdr_row+15}', CellIsRule(
        operator='equal', formula=['"Open"'], fill=fail_fill,
        font=Font(color=RED, bold=True)))
    ws.conditional_formatting.add(f'G{defect_hdr_row+1}:G{defect_hdr_row+15}', CellIsRule(
        operator='equal', formula=['"Fixed"'], fill=pass_fill,
        font=Font(color=GREEN, bold=True)))
    ws.conditional_formatting.add(f'G{defect_hdr_row+1}:G{defect_hdr_row+15}', CellIsRule(
        operator='equal', formula=['"Verified"'], fill=pass_fill,
        font=Font(color=GREEN, bold=True)))

    # ══════════════════════════════════════════
    # INSTRUCTIONS (after defects)
    # ══════════════════════════════════════════
    instr_row = defect_hdr_row + 18
    ws.merge_cells(f'B{instr_row}:{LAST_COL_LETTER}{instr_row}')
    ws.cell(row=instr_row, column=2).value = "HOW TO USE THIS WORKBOOK"
    ws.cell(row=instr_row, column=2).font = section_font

    instructions = [
        "1. Before testing: Go to Run Log and create a new row (date, tester, environment).",
        "2. Go to Summary sheet and fill PASS/FAIL/BLOCKED/SKIP per scenario in the next Run column.",
        "3. For step-level tracking: open individual scenario sheets (S-01, S-02, etc.).",
        "4. Dropdowns: PASS/FAIL/BLOCKED/SKIP for scenarios; \u2705/\u274c/\u23ed/\u26a0 for steps.",
        "5. Dashboard KPIs, charts, and heatmap auto-update from Summary formulas.",
        "6. Reference sheet has all test accounts, passwords, and Stripe test cards.",
        "7. Log defects in the Defect Tracker above \u2014 link to GitHub Issues.",
        "8. Reseed DEV data before each full regression run for clean state.",
        "9. Coverage sheet maps features to scenarios \u2014 update 'Last Tested' after runs.",
    ]
    for i, instr in enumerate(instructions):
        r = instr_row + 1 + i
        ws.merge_cells(f'B{r}:{LAST_COL_LETTER}{r}')
        ws.cell(row=r, column=2).value = instr
        ws.cell(row=r, column=2).font = Font(name='Calibri', size=11, color=DARK)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[r].height = 22

    # Hide gridlines for clean look
    ws.sheet_view.showGridLines = False


def create_run_log(wb):
    ws = wb.create_sheet("Run Log")
    ws.sheet_properties.tabColor = CORAL

    headers = [
        "Run #", "Date", "Tester", "Environment", "Seed Version",
        "Scenarios Attempted", "Passed", "Failed", "Blocked", "Skipped",
        "Pass Rate", "Duration", "Notes / Issues Found"
    ]

    widths = [8, 14, 18, 14, 14, 18, 10, 10, 10, 10, 12, 12, 45]

    # Title
    ws.merge_cells('A1:M1')
    ws.cell(row=1, column=1).value = "Test Run Log"
    ws.cell(row=1, column=1).font = Font(name='Calibri', bold=True, color=TEAL, size=16)
    ws.cell(row=1, column=1).fill = PatternFill(start_color=TEAL_LIGHT, end_color=TEAL_LIGHT, fill_type='solid')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    for i, (h, w) in enumerate(zip(headers, widths)):
        col = i + 1
        ws.column_dimensions[get_column_letter(col)].width = w
        cell = ws.cell(row=3, column=col)
        cell.value = h
    style_header_row(ws, 3, len(headers))

    # Pre-fill 20 rows with formulas
    for r in range(4, 24):
        ws.cell(row=r, column=1).value = r - 3
        ws.cell(row=r, column=1).alignment = center_align
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=1).font = body_font

        for c in range(2, 14):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = body_font
            cell.alignment = center_align if c < 12 else body_align

        # Pass Rate formula
        ws.cell(row=r, column=11).value = f'=IF(G{r}+H{r}>0,G{r}/(G{r}+H{r}),"")'
        ws.cell(row=r, column=11).number_format = '0%'

    # Date validation
    from openpyxl.worksheet.datavalidation import DataValidation
    env_dv = DataValidation(type="list", formula1='"DEV,STAGING,PROD"', showDropDown=False)
    ws.add_data_validation(env_dv)
    env_dv.add(f'D4:D23')

    # Alternating row colors
    for r in range(4, 24):
        if r % 2 == 0:
            for c in range(1, 14):
                ws.cell(row=r, column=c).fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid')

    ws.freeze_panes = 'A4'


def create_scenario_summary(wb):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = TEAL

    # Title
    ws.merge_cells('A1:N1')
    ws.cell(row=1, column=1).value = "Scenario Test Summary"
    ws.cell(row=1, column=1).font = Font(name='Calibri', bold=True, color=TEAL, size=16)
    ws.cell(row=1, column=1).fill = PatternFill(start_color=TEAL_LIGHT, end_color=TEAL_LIGHT, fill_type='solid')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    fixed_headers = ["ID", "Scenario Name", "Tier", "Roles", "Est. Time", "Steps", "Preconditions"]
    run_headers = [f"Run {i+1}" for i in range(NUM_RUN_COLUMNS)]
    latest_header = ["Latest Status"]
    notes_header = ["Bug Issues"]
    all_headers = fixed_headers + run_headers + latest_header + notes_header

    widths = [7, 35, 18, 28, 10, 8, 40] + [12]*NUM_RUN_COLUMNS + [14, 25]

    for i, (h, w) in enumerate(zip(all_headers, widths)):
        col = i + 1
        ws.column_dimensions[get_column_letter(col)].width = w
        ws.cell(row=2, column=col).value = h
    style_header_row(ws, 2, len(all_headers))

    # Scenario data rows
    for s_idx, s in enumerate(SCENARIOS):
        r = s_idx + 3
        ws.cell(row=r, column=1).value = s["id"]
        ws.cell(row=r, column=1).font = Font(name='Calibri', bold=True, color=TEAL)
        ws.cell(row=r, column=1).alignment = center_align
        ws.cell(row=r, column=1).border = thin_border

        ws.cell(row=r, column=2).value = s["name"]
        ws.cell(row=r, column=2).font = body_font
        ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=2).alignment = body_align

        ws.cell(row=r, column=3).value = s["tier"]
        ws.cell(row=r, column=3).font = body_font
        ws.cell(row=r, column=3).alignment = center_align
        ws.cell(row=r, column=3).border = thin_border

        ws.cell(row=r, column=4).value = s["roles"]
        ws.cell(row=r, column=4).font = body_font
        ws.cell(row=r, column=4).alignment = body_align
        ws.cell(row=r, column=4).border = thin_border

        ws.cell(row=r, column=5).value = s["est_time"]
        ws.cell(row=r, column=5).font = body_font
        ws.cell(row=r, column=5).alignment = center_align
        ws.cell(row=r, column=5).border = thin_border

        ws.cell(row=r, column=6).value = len(s["steps"])
        ws.cell(row=r, column=6).font = body_font
        ws.cell(row=r, column=6).alignment = center_align
        ws.cell(row=r, column=6).border = thin_border

        ws.cell(row=r, column=7).value = s["preconditions"]
        ws.cell(row=r, column=7).font = Font(name='Calibri', size=10, color=GRAY_TEXT)
        ws.cell(row=r, column=7).alignment = body_align
        ws.cell(row=r, column=7).border = thin_border

        # Run columns (empty with dropdowns)
        for run_col in range(8, 8 + NUM_RUN_COLUMNS):
            cell = ws.cell(row=r, column=run_col)
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = Font(name='Calibri', bold=True, size=11)

        # Latest Status formula (rightmost non-empty run)
        latest_col = 8 + NUM_RUN_COLUMNS
        # Build a nested formula to find the last non-empty run value
        run_cols_letters = [get_column_letter(c) for c in range(8, 8 + NUM_RUN_COLUMNS)]
        # Use a simpler approach: last non-blank from the run columns
        formula_parts = []
        for rc in reversed(run_cols_letters):
            formula_parts.append(f'{rc}{r}<>""')

        # Build nested IF: IF(last<>"", last, IF(second-last<>"", ...))
        latest_formula = '""'
        for rc in run_cols_letters:
            latest_formula = f'IF({rc}{r}<>"",{rc}{r},{latest_formula})'

        ws.cell(row=r, column=latest_col).value = f'={latest_formula}'
        ws.cell(row=r, column=latest_col).alignment = center_align
        ws.cell(row=r, column=latest_col).border = thin_border
        ws.cell(row=r, column=latest_col).font = Font(name='Calibri', bold=True, size=12)

        # Bug Issues column
        bug_col = latest_col + 1
        ws.cell(row=r, column=bug_col).border = thin_border
        ws.cell(row=r, column=bug_col).font = body_font
        ws.cell(row=r, column=bug_col).alignment = body_align

        # Alternating rows
        if s_idx % 2 == 0:
            for c in range(1, len(all_headers) + 1):
                cell = ws.cell(row=r, column=c)
                if not cell.fill or cell.fill.start_color.rgb == '00000000':
                    cell.fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid')

    # Add run status dropdowns
    last_data_row = len(SCENARIOS) + 2
    for run_col in range(8, 8 + NUM_RUN_COLUMNS):
        col_letter = get_column_letter(run_col)
        add_status_validation(ws, f'{col_letter}3:{col_letter}{last_data_row}')
        add_conditional_formatting(ws, f'{col_letter}3:{col_letter}{last_data_row}')

    # Latest Status conditional formatting
    latest_col_letter = get_column_letter(8 + NUM_RUN_COLUMNS)
    add_conditional_formatting(ws, f'{latest_col_letter}3:{latest_col_letter}{last_data_row}')

    # ── Totals Row ──
    totals_row = last_data_row + 1
    ws.cell(row=totals_row, column=1).value = "TOTALS"
    ws.cell(row=totals_row, column=1).font = Font(name='Calibri', bold=True, color=WHITE, size=11)
    ws.cell(row=totals_row, column=1).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
    ws.cell(row=totals_row, column=1).border = thin_border
    ws.cell(row=totals_row, column=1).alignment = center_align

    for c in range(2, len(all_headers) + 1):
        ws.cell(row=totals_row, column=c).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
        ws.cell(row=totals_row, column=c).border = thin_border

    # Totals in B (Pass count), C (Fail count), D (Blocked), E (Skip)
    latest_cl = get_column_letter(8 + NUM_RUN_COLUMNS)
    ws.cell(row=totals_row, column=2).value = f'=COUNTIF({latest_cl}3:{latest_cl}{last_data_row},"PASS")'
    ws.cell(row=totals_row, column=2).font = Font(name='Calibri', bold=True, color=GREEN, size=12)
    ws.cell(row=totals_row, column=2).alignment = center_align

    ws.cell(row=totals_row, column=3).value = f'=COUNTIF({latest_cl}3:{latest_cl}{last_data_row},"FAIL")'
    ws.cell(row=totals_row, column=3).font = Font(name='Calibri', bold=True, color=RED, size=12)
    ws.cell(row=totals_row, column=3).alignment = center_align

    ws.cell(row=totals_row, column=4).value = f'=COUNTIF({latest_cl}3:{latest_cl}{last_data_row},"BLOCKED")'
    ws.cell(row=totals_row, column=4).font = Font(name='Calibri', bold=True, color=AMBER, size=12)
    ws.cell(row=totals_row, column=4).alignment = center_align

    ws.cell(row=totals_row, column=5).value = f'=COUNTIF({latest_cl}3:{latest_cl}{last_data_row},"SKIP")'
    ws.cell(row=totals_row, column=5).font = Font(name='Calibri', bold=True, color=GRAY_TEXT, size=12)
    ws.cell(row=totals_row, column=5).alignment = center_align

    # Labels for totals
    labels_row = totals_row + 1
    for c, label in [(2, "Passed"), (3, "Failed"), (4, "Blocked"), (5, "Skipped")]:
        ws.cell(row=labels_row, column=c).value = label
        ws.cell(row=labels_row, column=c).font = Font(name='Calibri', size=9, color=GRAY_TEXT)
        ws.cell(row=labels_row, column=c).alignment = center_align

    ws.freeze_panes = 'H3'


def create_scenario_sheet(wb, scenario):
    sid = scenario["id"]
    # Sheet name limited to 31 chars
    sheet_name = f'{sid} {scenario["name"]}'[:31]
    ws = wb.create_sheet(sheet_name)

    # Color code by tier
    tier_colors = {
        "Foundation": "16A34A",
        "Core Marketplace": TEAL,
        "Communication": BLUE,
        "Business Ops": CORAL,
        "Admin & Platform": "7C3AED",
        "Edge Cases": AMBER,
        "Master Validation": RED,
    }
    ws.sheet_properties.tabColor = tier_colors.get(scenario["tier"], TEAL)

    # ── Header Section ──
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 50

    # Run columns
    for i in range(NUM_RUN_COLUMNS):
        ws.column_dimensions[get_column_letter(6 + i)].width = 10

    ws.column_dimensions[get_column_letter(6 + NUM_RUN_COLUMNS)].width = 35  # Notes

    # Title bar
    total_cols = 6 + NUM_RUN_COLUMNS + 1
    for col in range(1, total_cols + 1):
        ws.cell(row=1, column=col).fill = PatternFill(
            start_color=tier_colors.get(scenario["tier"], TEAL),
            end_color=tier_colors.get(scenario["tier"], TEAL),
            fill_type='solid'
        )

    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    ws.cell(row=1, column=1).value = f'{sid}: {scenario["name"]}'
    ws.cell(row=1, column=1).font = Font(name='Calibri', bold=True, color=WHITE, size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # Metadata
    meta = [
        ("Capability:", scenario["capability"]),
        ("Roles:", scenario["roles"]),
        ("Est. Time:", scenario["est_time"]),
        ("Preconditions:", scenario["preconditions"]),
    ]

    for i, (label, value) in enumerate(meta):
        r = 3 + i
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=1).font = Font(name='Calibri', bold=True, color=TEAL, size=11)
        ws.merge_cells(f'B{r}:{get_column_letter(total_cols)}{r}')
        ws.cell(row=r, column=2).value = value
        ws.cell(row=r, column=2).font = body_font

    # ── Step Table ──
    header_row = 8
    step_headers = ["#", "As (Role)", "Action", "Page", "Expected Result"]
    run_headers = [f"Run {i+1}" for i in range(NUM_RUN_COLUMNS)]
    notes_header = ["Notes / Bug #"]
    all_headers = step_headers + run_headers + notes_header

    for i, h in enumerate(all_headers):
        ws.cell(row=header_row, column=i+1).value = h
    style_header_row(ws, header_row, len(all_headers))

    # Step data
    for step_idx, (role, action, page, checkpoint) in enumerate(scenario["steps"]):
        r = header_row + 1 + step_idx

        ws.cell(row=r, column=1).value = step_idx + 1
        ws.cell(row=r, column=1).alignment = center_align
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=1).font = Font(name='Calibri', bold=True, color=TEAL)

        ws.cell(row=r, column=2).value = role
        ws.cell(row=r, column=2).font = Font(name='Calibri', bold=True, color=DARK, size=11)
        ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=2).alignment = body_align

        ws.cell(row=r, column=3).value = action
        ws.cell(row=r, column=3).font = body_font
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=3).alignment = body_align

        ws.cell(row=r, column=4).value = page
        ws.cell(row=r, column=4).font = Font(name='Calibri', size=10, color=GRAY_TEXT)
        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=4).alignment = center_align

        ws.cell(row=r, column=5).value = checkpoint
        ws.cell(row=r, column=5).font = body_font
        ws.cell(row=r, column=5).border = thin_border
        ws.cell(row=r, column=5).alignment = body_align

        # Run columns + notes
        for c in range(6, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = center_align
            cell.font = Font(name='Calibri', size=14)

        # Alternating rows
        if step_idx % 2 == 0:
            for c in range(1, total_cols + 1):
                ws.cell(row=r, column=c).fill = PatternFill(
                    start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid'
                )

        ws.row_dimensions[r].height = 30

    # Add step-level dropdowns and conditional formatting
    last_step_row = header_row + len(scenario["steps"])
    for run_col in range(6, 6 + NUM_RUN_COLUMNS):
        col_letter = get_column_letter(run_col)
        cell_range = f'{col_letter}{header_row+1}:{col_letter}{last_step_row}'
        add_step_status_validation(ws, cell_range)
        add_step_conditional_formatting(ws, cell_range)

    # ── Scenario Result Row ──
    result_row = last_step_row + 2
    ws.merge_cells(f'A{result_row}:E{result_row}')
    ws.cell(row=result_row, column=1).value = "SCENARIO RESULT"
    ws.cell(row=result_row, column=1).font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    ws.cell(row=result_row, column=1).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
    ws.cell(row=result_row, column=1).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=result_row, column=1).border = thin_border
    for c in range(2, 6):
        ws.cell(row=result_row, column=c).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
        ws.cell(row=result_row, column=c).border = thin_border

    for run_col in range(6, 6 + NUM_RUN_COLUMNS):
        cell = ws.cell(row=result_row, column=run_col)
        cell.border = thin_border
        cell.alignment = center_align
        cell.font = Font(name='Calibri', bold=True, size=14)
        cell.fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid')

    # Add PASS/FAIL dropdown to result row
    for run_col in range(6, 6 + NUM_RUN_COLUMNS):
        col_letter = get_column_letter(run_col)
        result_range = f'{col_letter}{result_row}'
        add_status_validation(ws, result_range)
        add_conditional_formatting(ws, result_range)

    # Notes for result
    ws.cell(row=result_row, column=6 + NUM_RUN_COLUMNS).border = thin_border
    ws.cell(row=result_row, column=6 + NUM_RUN_COLUMNS).fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid')

    ws.row_dimensions[result_row].height = 28
    ws.freeze_panes = f'F{header_row + 1}'


def create_reference_sheet(wb):
    ws = wb.create_sheet("Reference")
    ws.sheet_properties.tabColor = "6C757D"

    # Title
    ws.merge_cells('A1:G1')
    ws.cell(row=1, column=1).value = "Test Account Reference"
    ws.cell(row=1, column=1).font = Font(name='Calibri', bold=True, color=TEAL, size=16)
    ws.cell(row=1, column=1).fill = PatternFill(start_color=TEAL_LIGHT, end_color=TEAL_LIGHT, fill_type='solid')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:G2')
    ws.cell(row=2, column=1).value = "Universal Password: SeedTest2026!"
    ws.cell(row=2, column=1).font = Font(name='Calibri', bold=True, color=CORAL, size=13)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    # ── RAV Team ──
    row = 4
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1).value = "RAV TEAM ACCOUNTS"
    ws.cell(row=row, column=1).font = section_font

    row = 5
    headers = ["Alias", "Email", "Role", "Access", "Tier", "Notes", "Stripe ID"]
    widths = [14, 35, 14, 25, 10, 25, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        ws.column_dimensions[get_column_letter(i+1)].width = w
        ws.cell(row=row, column=i+1).value = h
    style_header_row(ws, row, len(headers))

    team_data = [
        ("RAV Owner", "dev-owner@rent-a-vacation.com", "rav_owner", "Full admin (all 19 tabs)", "\u2014", "Platform owner", "\u2014"),
        ("RAV Admin", "dev-admin@rent-a-vacation.com", "rav_admin", "Admin (all 19 tabs)", "\u2014", "Admin operations", "\u2014"),
        ("RAV Staff", "dev-staff@rent-a-vacation.com", "rav_staff", "Staff (10 operational tabs)", "\u2014", "Operational only", "\u2014"),
    ]
    for i, data in enumerate(team_data):
        r = row + 1 + i
        for c, val in enumerate(data):
            cell = style_data_cell(ws, r, c + 1)
            cell.value = val
        if i % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = PatternFill(start_color=TEAL_LIGHT, end_color=TEAL_LIGHT, fill_type='solid')

    # ── Property Owners ──
    row = 10
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1).value = "PROPERTY OWNER ACCOUNTS"
    ws.cell(row=row, column=1).font = section_font

    row = 11
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i+1).value = h
    style_header_row(ws, row, len(headers))

    owner_data = [
        ("Owner 1", "owner1@rent-a-vacation.com", "property_owner", "Owner dashboard, 10 listings", "Pro", "Alex Rivera, Hilton", "cus_test_owner1_pro"),
        ("Owner 2", "owner2@rent-a-vacation.com", "property_owner", "Owner dashboard, 25 listings", "Business", "Maria Chen, Marriott", "cus_test_owner2_biz"),
        ("Owner 3", "owner3@rent-a-vacation.com", "property_owner", "Owner dashboard, 3 listings", "Free", "James Thompson, Disney", "\u2014"),
        ("Owner 4", "owner4@rent-a-vacation.com", "property_owner", "Owner dashboard, 3 listings", "Free", "Priya Patel, Wyndham", "\u2014"),
        ("Owner 5", "owner5@rent-a-vacation.com", "property_owner", "Owner dashboard, 3 listings", "Free", "Robert Kim, Bluegreen", "\u2014"),
    ]
    for i, data in enumerate(owner_data):
        r = row + 1 + i
        for c, val in enumerate(data):
            cell = style_data_cell(ws, r, c + 1)
            cell.value = val
        if i % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid')

    # ── Key Renters ──
    row = 18
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1).value = "KEY RENTER ACCOUNTS"
    ws.cell(row=row, column=1).font = section_font

    row = 19
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i+1).value = h
    style_header_row(ws, row, len(headers))

    renter_data = [
        ("Renter 1", "renter001@rent-a-vacation.com", "renter", "Marketplace + My Trips", "Plus", "Sophia Martinez", "cus_test_renter1_plus"),
        ("Renter 2", "renter002@rent-a-vacation.com", "renter", "Marketplace + My Trips", "Plus", "Liam Johnson", "cus_test_renter2_plus"),
        ("Renter 3", "renter003@rent-a-vacation.com", "renter", "Marketplace + My Trips", "Premium", "Olivia Williams", "cus_test_renter3_prem"),
        ("Renter 4", "renter004@rent-a-vacation.com", "renter", "Marketplace + My Trips", "Free", "\u2014", "\u2014"),
        ("Renter 50", "renter050@rent-a-vacation.com", "renter", "Marketplace + My Trips", "Free", "Expendable (GDPR test)", "\u2014"),
    ]
    for i, data in enumerate(renter_data):
        r = row + 1 + i
        for c, val in enumerate(data):
            cell = style_data_cell(ws, r, c + 1)
            cell.value = val
        if i % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid')

    # ── Stripe Test Cards ──
    row = 26
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1).value = "STRIPE TEST CARDS"
    ws.cell(row=row, column=1).font = section_font

    row = 27
    card_headers = ["Scenario", "Card Number", "CVC", "Expiry", "ZIP", "Expected", "Used In"]
    for i, h in enumerate(card_headers):
        ws.cell(row=row, column=i+1).value = h
    style_header_row(ws, row, len(card_headers))

    card_data = [
        ("Success (default)", "4242 4242 4242 4242", "123", "12/28", "32256", "Payment completes", "S-04, S-05, S-06, S-30"),
        ("Visa debit", "4000 0566 5566 5556", "123", "12/28", "32256", "Payment completes", "S-19"),
        ("Insufficient funds", "4000 0000 0000 9995", "123", "12/28", "32256", "Decline: insufficient", "S-19"),
        ("Generic decline", "4000 0000 0000 0002", "123", "12/28", "32256", "Decline: generic", "S-19"),
        ("Expired card", "4000 0000 0000 0069", "123", "12/28", "32256", "Decline: expired", "S-19"),
        ("3D Secure (pass)", "4000 0027 6000 3184", "123", "12/28", "32256", "3DS popup \u2192 success", "S-19"),
        ("3D Secure (fail)", "4000 0084 0000 1629", "123", "12/28", "32256", "3DS popup \u2192 decline", "S-19"),
    ]
    for i, data in enumerate(card_data):
        r = row + 1 + i
        for c, val in enumerate(data):
            cell = style_data_cell(ws, r, c + 1)
            cell.value = val
        if i % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid')

    # ── Environment Info ──
    row = 36
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1).value = "ENVIRONMENT"
    ws.cell(row=row, column=1).font = section_font

    env_data = [
        ("DEV URL", "https://dev.rent-a-vacation.com"),
        ("PROD URL", "https://rent-a-vacation.com (DO NOT TEST)"),
        ("Stripe Dashboard", "https://dashboard.stripe.com (toggle Test mode)"),
        ("GitHub Issues", "https://github.com/rent-a-vacation/rav-website/issues"),
        ("Supabase DEV", "oukbxqnlxnkainnligfz"),
        ("Reseed Command", "Admin \u2192 Dev Tools \u2192 Reset & Reseed DEV"),
    ]
    for i, (label, value) in enumerate(env_data):
        r = row + 1 + i
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=1).font = Font(name='Calibri', bold=True, color=TEAL, size=11)
        ws.cell(row=r, column=1).border = thin_border
        ws.merge_cells(f'B{r}:G{r}')
        ws.cell(row=r, column=2).value = value
        ws.cell(row=r, column=2).font = body_font
        ws.cell(row=r, column=2).border = thin_border

    ws.freeze_panes = 'A2'


def create_coverage_matrix(wb):
    ws = wb.create_sheet("Coverage")
    ws.sheet_properties.tabColor = BLUE

    # Title
    ws.merge_cells('A1:E1')
    ws.cell(row=1, column=1).value = "Feature Coverage Matrix"
    ws.cell(row=1, column=1).font = Font(name='Calibri', bold=True, color=TEAL, size=16)
    ws.cell(row=1, column=1).fill = PatternFill(start_color=TEAL_LIGHT, end_color=TEAL_LIGHT, fill_type='solid')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    headers = ["Feature Area", "Scenarios Covering It", "# Scenarios", "Last Tested", "Status"]
    widths = [30, 50, 14, 16, 14]
    for i, (h, w) in enumerate(zip(headers, widths)):
        ws.column_dimensions[get_column_letter(i+1)].width = w
        ws.cell(row=3, column=i+1).value = h
    style_header_row(ws, 3, len(headers))

    coverage_data = [
        ("User Registration & Approval", "S-01, S-15", 2),
        ("Role Upgrade & Verification", "S-02", 1),
        ("Property Listing (Create/Edit)", "S-03, S-21, S-24", 3),
        ("Direct Booking & Payment", "S-04, S-19, S-30", 3),
        ("Bidding & Counter-Offers", "S-05, S-30", 2),
        ("Travel Requests & Proposals", "S-06", 1),
        ("Date Proposals (Flexible)", "S-23", 1),
        ("Pre-Booking Messaging", "S-07", 1),
        ("Booking Messaging (Realtime)", "S-08, S-30", 2),
        ("Renter Cancellation", "S-09", 1),
        ("Owner Cancellation", "S-10", 1),
        ("Disputes & Evidence", "S-11", 1),
        ("Owner Earnings & Payouts", "S-12, S-30", 2),
        ("Membership Upgrade (Owner)", "S-13, S-21", 2),
        ("Membership Upgrade (Renter)", "S-14", 1),
        ("Referral Program", "S-15", 1),
        ("Search, Filter, Sort", "S-16", 1),
        ("Compare Listings", "S-16", 1),
        ("Saved Searches", "S-16", 1),
        ("Destinations", "S-16, S-20", 2),
        ("Notifications & Preferences", "S-17", 1),
        ("Admin Operations (19 tabs)", "S-18", 1),
        ("Admin Property Editing", "S-24", 1),
        ("Admin Listing Editing", "S-24", 1),
        ("Executive Dashboard", "S-25", 1),
        ("Voice Search & Quotas", "S-26", 1),
        ("API Keys & Developer Portal", "S-27", 1),
        ("iCal Export", "S-28", 1),
        ("GDPR Export & Deletion", "S-29", 1),
        ("RAV Smart Tools (5 tools)", "S-22", 1),
        ("Public Pages & Route Guards", "S-20", 1),
        ("Stripe Payment Edge Cases", "S-19", 1),
        ("Staff Role Boundaries", "S-18, S-25", 2),
        ("Owner Confirmation Window", "S-04, S-30", 2),
        ("Escrow Management", "S-04, S-12, S-30", 3),
        ("Full Marketplace Lifecycle", "S-30", 1),
    ]

    for i, (feature, scenarios, count) in enumerate(coverage_data):
        r = 4 + i
        ws.cell(row=r, column=1).value = feature
        ws.cell(row=r, column=1).font = body_font
        ws.cell(row=r, column=1).border = thin_border

        ws.cell(row=r, column=2).value = scenarios
        ws.cell(row=r, column=2).font = Font(name='Calibri', size=11, color=TEAL)
        ws.cell(row=r, column=2).border = thin_border

        ws.cell(row=r, column=3).value = count
        ws.cell(row=r, column=3).alignment = center_align
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=3).font = Font(name='Calibri', bold=True, color=TEAL)

        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=4).alignment = center_align
        ws.cell(row=r, column=4).font = body_font
        ws.cell(row=r, column=4).number_format = 'YYYY-MM-DD'

        ws.cell(row=r, column=5).border = thin_border
        ws.cell(row=r, column=5).alignment = center_align

        if i % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type='solid')

    # Status validation
    from openpyxl.worksheet.datavalidation import DataValidation
    last_r = 3 + len(coverage_data)
    dv = DataValidation(type="list", formula1='"Covered,Partial,Not Tested,Needs Update"', showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f'E4:E{last_r}')

    # Conditional formatting for status
    ws.conditional_formatting.add(f'E4:E{last_r}', CellIsRule(
        operator='equal', formula=['"Covered"'], fill=pass_fill,
        font=Font(color=GREEN, bold=True)
    ))
    ws.conditional_formatting.add(f'E4:E{last_r}', CellIsRule(
        operator='equal', formula=['"Not Tested"'], fill=fail_fill,
        font=Font(color=RED, bold=True)
    ))
    ws.conditional_formatting.add(f'E4:E{last_r}', CellIsRule(
        operator='equal', formula=['"Partial"'], fill=blocked_fill,
        font=Font(color=AMBER, bold=True)
    ))

    ws.freeze_panes = 'A4'


# ── MAIN ──
if __name__ == "__main__":
    print("Generating RAV QA Scenario Test Workbook...")
    wb = create_workbook()

    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'docs', 'testing')
    output_path = os.path.join(output_dir, 'RAV-QA-Scenario-Tests.xlsx')
    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Total sheets: {len(wb.sheetnames)}")
